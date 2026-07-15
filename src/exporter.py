"""
Exporter - Export hasil scraping ke Excel dengan sheet lengkap.
Includes: Summary, Monitoring_Result, Field_Status, Debug_Log, Failed_Posts
"""
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from src.parser import AccountRow, ScrapeRow

def export_debug_log_csv(output_path: Path) -> Optional[Path]:
    """Placeholder for debug log export."""
    return None

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SECTION_FONT = Font(bold=True, size=11)
SUCCESS_FILL = PatternFill("solid", fgColor="C6EFCE")
WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)

HEADER_ROW = 6
TITLE_ROW = 8
DATA_START_ROW = 9

def style_header(sheet, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_body(sheet, min_row: int, max_row: int, min_col: int, max_col: int):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def apply_status_color(cell, status: str):
    """Apply color berdasarkan status scraping."""
    if not status:
        return

    status_lower = status.lower()

    if "success" in status_lower and "partial" not in status_lower:
        cell.fill = SUCCESS_FILL
    elif "partial" in status_lower or "null" in status_lower or "review" in status_lower:
        cell.fill = WARNING_FILL
    elif any(x in status_lower for x in ["fail", "error", "login", "not_found", "rate", "timeout"]):
        cell.fill = ERROR_FILL


def set_column_widths(sheet, widths: Dict[str, int]):
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width


def format_datetime_cell(sheet, row: int, col: int):
    """Format cell sebagai datetime."""
    sheet.cell(row=row, column=col).number_format = "dd mmmm yyyy hh:mm"

def build_output(
    accounts: List[AccountRow],
    rows: List[ScrapeRow],
    selected_fields: List[str],
    only_period: bool = False,
    include_raw: bool = True,
    include_extended: bool = False,
) -> bytes:
    """
    Build output workbook dengan multiple sheets:
    - Summary
    - Monitoring_Result
    - Raw_Scraping (if include_raw)
    - Failed_Posts

    Args:
        include_extended: If True, include extended metrics columns in exports.
                         Default False to maintain backward compatibility.
    """
    workbook = Workbook()

    # Sheet 1: Summary
    build_summary_sheet(workbook, accounts, rows)

    # Sheet 2: Monitoring Result
    build_monitoring_sheet(workbook, accounts, rows, selected_fields, only_period, include_extended)

    # Sheet 3: Raw Scraping
    if include_raw:
        build_raw_sheet(workbook, rows)

    # Sheet 4: Failed Posts
    failed_rows = [r for r in rows if r.status_scraping and any(
        x in r.status_scraping.lower()
        for x in ["fail", "error", "login", "not_found", "rate", "null", "timeout"]
    )]
    if failed_rows:
        build_failed_sheet(workbook, failed_rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_summary_sheet(workbook: Workbook, accounts: List[AccountRow], rows: List[ScrapeRow]):
    """Build summary sheet."""
    sheet = workbook.active
    sheet.title = "Summary"

    # Title
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "MAYZ SCRAPING SUMMARY"
    sheet["A1"].font = Font(bold=True, size=14)

    sheet.merge_cells("A2:D2")
    sheet["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sheet["A2"].font = Font(size=10, italic=True)

    # Stats
    stats_start_row = 4

    total_accounts = len(accounts)
    total_posts = len(rows)
    posts_with_url = len([r for r in rows if r.post_url])
    posts_with_caption = len([r for r in rows if r.caption])
    posts_with_timestamp = len([r for r in rows if r.tanggal_postingan])
    posts_with_likes = len([r for r in rows if r.like_count is not None])
    posts_with_comments = len([r for r in rows if r.comment_count is not None])

    total_likes = sum(r.like_count or 0 for r in rows)
    total_comments = sum(r.comment_count or 0 for r in rows)
    total_engagement = sum(r.total_engagement or 0 for r in rows)

    # Status breakdown
    status_counts = {}
    for r in rows:
        status = r.status_scraping or "Unknown"
        status_counts[status] = status_counts.get(status, 0) + 1

    stats = [
        ("Total Akun Diproses", total_accounts),
        ("Total Postingan", total_posts),
        ("Postingan dengan URL", posts_with_url),
        ("Postingan dengan Caption", posts_with_caption),
        ("Postingan dengan Timestamp", posts_with_timestamp),
        ("Postingan dengan Like", posts_with_likes),
        ("Postingan dengan Comment", posts_with_comments),
        ("", ""),
        ("Total Like", total_likes),
        ("Total Comment", total_comments),
        ("Total Engagement", total_engagement),
    ]

    for i, (label, value) in enumerate(stats, start=stats_start_row):
        sheet.cell(row=i, column=1).value = label
        sheet.cell(row=i, column=1).font = Font(bold=True)
        sheet.cell(row=i, column=2).value = value

    # Status breakdown
    status_start = stats_start_row + len(stats) + 2
    sheet.cell(row=status_start, column=1).value = "Status Breakdown"
    sheet.cell(row=status_start, column=1).font = Font(bold=True)

    for i, (status, count) in enumerate(sorted(status_counts.items()), start=status_start + 1):
        sheet.cell(row=i, column=1).value = status
        sheet.cell(row=i, column=2).value = count

    # Success rate
    if posts_with_url > 0:
        caption_rate = (posts_with_caption / posts_with_url) * 100
        rate_row = status_start + len(status_counts) + 2
        sheet.cell(row=rate_row, column=1).value = "Caption Recovery Rate"
        sheet.cell(row=rate_row, column=1).font = Font(bold=True)
        sheet.cell(row=rate_row, column=2).value = f"{caption_rate:.1f}%"

    set_column_widths(sheet, {"A": 30, "B": 20, "C": 15, "D": 15})


def build_monitoring_sheet(
    workbook: Workbook,
    accounts: List[AccountRow],
    rows: List[ScrapeRow],
    selected_fields: List[str],
    only_period: bool,
    include_extended: bool = False
):
    """Build main monitoring result sheet.

    Args:
        include_extended: If True, include extended metrics columns (view/play/share/save/reach).
                         Default False to maintain existing layout.
    """
    sheet = workbook.create_sheet("Monitoring_Result")

    # Base headers
    base_headers = [
        "No.",
        "Nama Kanwil",
        "URL Akun",
        "Tanggal Postingan",
        "Caption",
        "Link",
        "Jenis Media",
        "Like Count",
        "Comment Count",
        "Total Engagement",
        "Status Scraping",
        "Status Periode",
        "Catatan",
    ]

    # Extended headers (optional)
    extended_headers = [
        "View Count",
        "Play Count",
        "Share Count",
        "Save Count",
        "Reach Count",
    ]

    all_headers = base_headers + (extended_headers if include_extended else [])

    last_col = len(all_headers)
    last_col_letter = get_column_letter(last_col)

    # Title
    sheet.merge_cells(f"A1:{last_col_letter}1")
    sheet["A1"] = "DAFTAR LINK PELAPORAN EKSIS VERTIKAL"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells(f"A2:{last_col_letter}2")
    sheet["A2"] = "DIREKTORAT JENDERAL PERBENDAHARAAN"
    sheet["A2"].font = Font(bold=True, size=12)
    sheet["A2"].alignment = Alignment(horizontal="center")

    sheet["A4"] = datetime.now().strftime("%B %Y")
    sheet["A4"].font = Font(bold=True)

    # Headers
    for i, header in enumerate(all_headers, start=1):
        sheet.cell(row=HEADER_ROW, column=i).value = header
    style_header(sheet, HEADER_ROW, 1, last_col)

    sheet.cell(row=TITLE_ROW, column=1).value = "Hasil Scraping Public Instagram"
    sheet.cell(row=TITLE_ROW, column=1).font = SECTION_FONT

    # Group by account
    grouped = {}
    for item in rows:
        grouped.setdefault(item.nama_kanwil, []).append(item)

    output_row = DATA_START_ROW
    number = 1

    for account in accounts:
        candidates = [item for item in grouped.get(account.nama_kanwil, []) if item.post_url]

        if only_period:
            candidates = [item for item in candidates if item.status_periode == "Masuk Periode"]

        for item in candidates:
            values = [
                number,
                account.nama_kanwil,
                account.url_akun,
                item.tanggal_postingan or "",
                item.caption or "",
                item.post_url,
                item.media_type or "unknown",
                item.like_count if item.like_count is not None else "",
                item.comment_count if item.comment_count is not None else "",
                item.total_engagement if item.total_engagement is not None else "",
                item.status_scraping or "",
                item.status_periode or "",
                item.catatan or "",
            ]

            # Add extended metrics if enabled
            if include_extended:
                values.extend([
                    item.get_extended_display("view_count"),
                    item.get_extended_display("play_count"),
                    item.get_extended_display("share_count"),
                    item.get_extended_display("save_count"),
                    item.get_extended_display("reach_count"),
                ])

            for col, value in enumerate(values, start=1):
                cell = sheet.cell(row=output_row, column=col)
                cell.value = value

                # Status column (column 11)
                if col == 11:
                    apply_status_color(cell, value)

            if item.post_url:
                sheet.cell(row=output_row, column=6).hyperlink = item.post_url
                sheet.cell(row=output_row, column=6).style = "Hyperlink"

            format_datetime_cell(sheet, output_row, 4)

            number += 1
            output_row += 1

    if output_row == DATA_START_ROW:
        sheet.cell(row=DATA_START_ROW, column=1).value = "Tidak ada data postingan yang berhasil dipetakan."

    style_body(sheet, DATA_START_ROW, max(output_row - 1, DATA_START_ROW), 1, last_col)

    # Standard column widths
    widths = {
        "A": 6,
        "B": 35,
        "C": 45,
        "D": 20,
        "E": 70,
        "F": 60,
        "G": 15,
        "H": 12,
        "I": 14,
        "J": 16,
        "K": 20,
        "L": 18,
        "M": 50,
    }

    # Extended column widths (if enabled)
    if include_extended:
        widths.update({
            "N": 14,  # View Count
            "O": 14,  # Play Count
            "P": 14,  # Share Count
            "Q": 14,  # Save Count
            "R": 14,  # Reach Count
        })

    set_column_widths(sheet, widths)

    sheet.freeze_panes = f"A{DATA_START_ROW}"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{max(output_row - 1, DATA_START_ROW)}"


def build_raw_sheet(workbook: Workbook, rows: List[ScrapeRow]):
    """Build raw scraping data sheet."""
    sheet = workbook.create_sheet("Raw_Scraping")

    headers = [
        "nama_kanwil",
        "url_akun",
        "post_url",
        "shortcode",
        "tanggal_postingan",
        "media_type",
        "caption",
        "like_count",
        "comment_count",
        "total_engagement",
        "view_count",
        "play_count",
        "share_count",
        "save_count",
        "reach_count",
        "status_periode",
        "status_scraping",
        "catatan",
    ]

    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i).value = header
    style_header(sheet, 1, 1, len(headers))

    for row_idx, item in enumerate(rows, start=2):
        values = [
            item.nama_kanwil,
            item.url_akun,
            item.post_url,
            item.shortcode,
            item.tanggal_postingan,
            item.media_type,
            item.caption,
            item.like_count if item.like_count is not None else "",
            item.comment_count if item.comment_count is not None else "",
            item.total_engagement if item.total_engagement is not None else "",
            item.get_extended_display("view_count"),
            item.get_extended_display("play_count"),
            item.get_extended_display("share_count"),
            item.get_extended_display("save_count"),
            item.get_extended_display("reach_count"),
            item.status_periode,
            item.status_scraping,
            item.catatan,
        ]

        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col)
            cell.value = value

            if col == 17:  # status_scraping column
                apply_status_color(cell, value)

        if item.tanggal_postingan:
            format_datetime_cell(sheet, row_idx, 5)

    style_body(sheet, 2, max(len(rows) + 1, 2), 1, len(headers))

    set_column_widths(sheet, {
        "A": 40,
        "B": 50,
        "C": 65,
        "D": 22,
        "E": 22,
        "F": 20,
        "G": 80,
        "H": 12,
        "I": 14,
        "J": 16,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 14,
        "P": 18,
        "Q": 20,
        "R": 60,
    })

    sheet.freeze_panes = "A2"


def build_failed_sheet(workbook: Workbook, failed_rows: List[ScrapeRow]):
    """Build failed posts sheet untuk review."""
    sheet = workbook.create_sheet("Failed_Posts")

    headers = [
        "No.",
        "Nama Kanwil",
        "post_url",
        "shortcode",
        "status_scraping",
        "catatan",
        "timestamp_tersedia",
        "caption_tersedia",
    ]

    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i).value = header
    style_header(sheet, 1, 1, len(headers))

    for row_idx, item in enumerate(failed_rows, start=2):
        values = [
            row_idx - 1,
            item.nama_kanwil,
            item.post_url,
            item.shortcode,
            item.status_scraping,
            item.catatan,
            "Ya" if item.tanggal_postingan else "Tidak",
            "Ya" if item.caption else "Tidak",
        ]

        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col)
            cell.value = value

            if col == 5:
                apply_status_color(cell, value)

        if item.post_url:
            sheet.cell(row=row_idx, column=3).hyperlink = item.post_url

    style_body(sheet, 2, len(failed_rows) + 1, 1, len(headers))

    set_column_widths(sheet, {
        "A": 6,
        "B": 40,
        "C": 60,
        "D": 22,
        "E": 22,
        "F": 60,
        "G": 18,
        "H": 18,
    })

    sheet.freeze_panes = "A2"


def save_output(content: bytes, export_dir: Path) -> Path:
    """Save output to file."""
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"mayz_output_{timestamp}.xlsx"
    path.write_bytes(content)
    return path


def save_output_with_debug(
    content: bytes,
    export_dir: Path,
    export_debug: bool = True
) -> tuple:
    """
    Save output AND export debug log.
    Returns: (output_path, debug_csv_path)
    """
    output_path = save_output(content, export_dir)

    debug_csv_path = None
    if export_debug:
        try:
            debug_csv_path = export_debug_log_csv(export_dir / "debug_log.csv")
        except Exception:
            pass  # Ignore debug export errors

    return output_path, debug_csv_path

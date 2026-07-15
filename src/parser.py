import re
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from typing import Any, List, Optional, Tuple
from urllib.parse import urlparse
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from src.config import SHEET_SOURCE

TZ_WIB = ZoneInfo("Asia/Jakarta")


@dataclass
class FieldEvidence:
    """Evidence untuk satu field yang di-extract."""
    field_name: str
    value: Any = None
    source: Optional[str] = None
    selector_used: Optional[str] = None
    extraction_method: Optional[str] = None
    status: str = "NULL"
    null_reason: Optional[str] = None
    attempted_selectors: List[str] = field(default_factory=list)

    def is_ok(self) -> bool:
        return self.value is not None and self.status == "OK"


@dataclass
class ExtendedMetrics:
    """
    Extended metrics yang diambil dari Instagram post.
    Field-field ini umumnya tidak tersedia dari public scraping
    dan mungkin memerlukan akses insights/private API.
    """
    # Standard engagement metrics
    like_count: Optional[int] = None
    comment_count: Optional[int] = None

    # Video/Reels metrics
    view_count: Optional[int] = None  # General view count
    play_count: Optional[int] = None  # Reels play count

    # Extended engagement (may require insights)
    share_count: Optional[int] = None
    save_count: Optional[int] = None
    reach_count: Optional[int] = None

    # Evidence tracking per field
    like_evidence: Optional['FieldEvidence'] = None
    comment_evidence: Optional['FieldEvidence'] = None
    view_evidence: Optional['FieldEvidence'] = None
    play_evidence: Optional['FieldEvidence'] = None
    share_evidence: Optional['FieldEvidence'] = None
    save_evidence: Optional['FieldEvidence'] = None
    reach_evidence: Optional['FieldEvidence'] = None

    # Notes
    insight_available: bool = False  # True if insights data was found
    insight_note: str = ""  # Note about why insights not available

    def total_engagement(self) -> int:
        """Calculate total engagement (like + comment only)."""
        total = 0
        if self.like_count:
            total += self.like_count
        if self.comment_count:
            total += self.comment_count
        return total if total > 0 else 0

    def extended_engagement(self) -> int:
        """
        Calculate extended engagement including shares and saves.
        Note: This is optional; don't use for main KPI.
        """
        total = self.total_engagement()
        if self.share_count:
            total += self.share_count
        if self.save_count:
            total += self.save_count
        return total if total > 0 else 0

    def get_evidence(self, field_name: str) -> Optional['FieldEvidence']:
        """Get evidence for a specific field."""
        evidence_map = {
            "like_count": self.like_evidence,
            "comment_count": self.comment_evidence,
            "view_count": self.view_evidence,
            "play_count": self.play_evidence,
            "share_count": self.share_evidence,
            "save_count": self.save_evidence,
            "reach_count": self.reach_evidence,
        }
        return evidence_map.get(field_name)


def parse_extended_metrics(page) -> ExtendedMetrics:
    """
    Parse extended metrics from page.
    Best-effort approach - returns None values for unavailable metrics.
    """
    metrics = ExtendedMetrics()

    # Try to extract from script JSON data first
    script_data = _extract_script_json(page)

    # Try og:description for engagement metrics
    og_description = _get_meta_content(page, "og:description") or ""

    # Extract view/play counts
    metrics.view_count, metrics.view_evidence = _extract_from_script_or_regex(
        page, script_data, "view_count", og_description,
        patterns=[
            r'"video_view_count"\s*:\s*(\d+)',
            r'"view_count"\s*:\s*(\d+)',
        ],
        regex_patterns=[
            r"([\d,.]+[KMRB]?)\s+(?:views?|tayangan)",
        ]
    )

    # Reels-specific play count
    metrics.play_count, metrics.play_evidence = _extract_from_script_or_regex(
        page, script_data, "play_count", og_description,
        patterns=[
            r'"play_count"\s*:\s*(\d+)',
        ],
        regex_patterns=[
            r"([\d,.]+[KMRB]?)\s+(?:plays?|pemutaran)",
        ]
    )

    # Like count (priority - try multiple sources)
    metrics.like_count, metrics.like_evidence = _extract_like_count(page, script_data, og_description)

    # Comment count (priority - try multiple sources)
    metrics.comment_count, metrics.comment_evidence = _extract_comment_count(page, script_data, og_description)

    # Share count (usually requires insights)
    metrics.share_count, metrics.share_evidence = _extract_from_script_or_regex(
        page, script_data, "share_count", og_description,
        patterns=[r'"share_count"\s*:\s*(\d+)'],
        regex_patterns=[r"([\d,.]+[KMRB]?)\s+(?:shares?|dibagikan)"],
        is_insight=True
    )

    # Save count (usually requires insights)
    metrics.save_count, metrics.save_evidence = _extract_from_script_or_regex(
        page, script_data, "save_count", og_description,
        patterns=[r'"save_count"\s*:\s*(\d+)'],
        regex_patterns=[r"([\d,.]+[KMRB]?)\s+(?:saves?|disimpan)"],
        is_insight=True
    )

    # Reach count (usually requires insights)
    metrics.reach_count, metrics.reach_evidence = _extract_from_script_or_regex(
        page, script_data, "reach_count", og_description,
        patterns=[r'"reach_count"\s*:\s*(\d+)'],
        regex_patterns=[r"([\d,.]+[KMRB]?)\s+(?:reach|jangkauan)"],
        is_insight=True
    )

    # Check if any insight data was found
    metrics.insight_available = any([
        metrics.share_count is not None,
        metrics.save_count is not None,
        metrics.reach_count is not None,
    ])

    if not metrics.insight_available:
        metrics.insight_note = "Metrik insight (share/save/reach) tidak tersedia dari public scraping. Diperlukan akses Instagram Insights atau API."

    return metrics


def _extract_script_json(page) -> str:
    """Extract JSON data from script tags."""
    try:
        # Try to find _sharedData in scripts
        script_locator = page.locator('script:not([src])')
        count = script_locator.count()
        for i in range(min(count, 20)):  # Limit search
            text = script_locator.nth(i).inner_text()
            if '"like_count"' in text or '"comments_count"' in text or '"view_count"' in text:
                return text
    except Exception:
        pass
    return ""


def _get_meta_content(page, property_name: str) -> Optional[str]:
    """Get meta content by property name."""
    try:
        locator = page.locator(f'meta[property="{property_name}"]')
        if locator.count() > 0:
            return locator.first.get_attribute("content")
    except Exception:
        pass
    return None


def _extract_from_script_or_regex(
    page,
    script_data: str,
    field_name: str,
    og_description: str,
    patterns: List[str],
    regex_patterns: List[str],
    is_insight: bool = False
) -> Tuple[Optional[int], Optional[FieldEvidence]]:
    """
    Extract numeric value from script JSON or regex fallback.
    Returns (value, evidence).
    """
    evidence = FieldEvidence(field_name=field_name)

    # Try script JSON patterns first
    for pattern in patterns:
        match = re.search(pattern, script_data)
        if match:
            value = parse_number_token(match.group(1))
            if value is not None:
                evidence.value = value
                evidence.source = "script_json"
                evidence.status = "OK"
                evidence.extraction_method = f"regex:{pattern}"
                return value, evidence

    # Try og:description regex fallback
    for pattern in regex_patterns:
        match = re.search(pattern, og_description.lower())
        if match:
            value = parse_number_token(match.group(1))
            if value is not None:
                evidence.value = value
                evidence.source = "og_description"
                evidence.status = "OK"
                evidence.extraction_method = f"regex:{pattern}"
                return value, evidence

    # Not found - mark as null (not failed)
    evidence.status = "NULL"
    if is_insight:
        evidence.null_reason = "Insight metric not available from public scraping"
    else:
        evidence.null_reason = "Metric not found in page content"

    return None, evidence


def _extract_like_count(page, script_data: str, og_description: str) -> Tuple[Optional[int], Optional[FieldEvidence]]:
    """Extract like count with multiple fallbacks."""
    evidence = FieldEvidence(field_name="like_count")

    # Priority 1: Script JSON
    match = re.search(r'"like_count"\s*:\s*(\d+)', script_data)
    if match:
        value = int(match.group(1))
        evidence.value = value
        evidence.source = "script_json"
        evidence.status = "OK"
        evidence.extraction_method = 'regex:"like_count"'
        return value, evidence

    # Priority 2: og:description pattern
    match = re.search(r"([\d,.]+[KMRB]?)\s+(?:likes?|suka)", og_description.lower())
    if match:
        value = parse_number_token(match.group(1))
        if value:
            evidence.value = value
            evidence.source = "og_description"
            evidence.status = "OK"
            evidence.extraction_method = "regex_like_pattern"
            return value, evidence

    # Priority 3: Try span with class pattern (Instagram's new structure)
    try:
        span_locator = page.locator('span[class*="x1ypdohk"]')
        if span_locator.count() > 0:
            text = span_locator.first.inner_text()
            value = parse_number_token(text)
            if value and value > 0:
                evidence.value = value
                evidence.source = "dom_span"
                evidence.selector_used = 'span[class*="x1ypdohk"]'
                evidence.status = "OK"
                evidence.extraction_method = "dom_text"
                return value, evidence
    except Exception:
        pass

    # Priority 4: Try SVG aria-label for likes
    try:
        like_svg = page.locator('svg[aria-label="Like"]')
        if like_svg.count() > 0:
            aria_label = like_svg.first.get_attribute("aria-label")
            if aria_label:
                # Extract number from "12 likes" or "12 suka"
                match = re.search(r"([\d,.]+[KMRB]?)", aria_label)
                if match:
                    value = parse_number_token(match.group(1))
                    if value:
                        evidence.value = value
                        evidence.source = "aria_label"
                        evidence.status = "OK"
                        evidence.extraction_method = 'svg[aria-label="Like"]'
                        return value, evidence
    except Exception:
        pass

    # Not found
    evidence.status = "NULL"
    evidence.null_reason = "Like count not available from public page"
    return None, evidence


def _extract_comment_count(page, script_data: str, og_description: str) -> Tuple[Optional[int], Optional[FieldEvidence]]:
    """Extract comment count with multiple fallbacks."""
    evidence = FieldEvidence(field_name="comment_count")

    # Priority 1: Script JSON
    match = re.search(r'"comments_count"\s*:\s*(\d+)', script_data)
    if match:
        value = int(match.group(1))
        evidence.value = value
        evidence.source = "script_json"
        evidence.status = "OK"
        evidence.extraction_method = 'regex:"comments_count"'
        return value, evidence

    # Priority 2: og:description pattern
    match = re.search(r"([\d,.]+[KMRB]?)\s+(?:comments?|komentar)", og_description.lower())
    if match:
        value = parse_number_token(match.group(1))
        if value:
            evidence.value = value
            evidence.source = "og_description"
            evidence.status = "OK"
            evidence.extraction_method = "regex_comment_pattern"
            return value, evidence

    # Priority 3: Check if comment section exists (indicates post has comments)
    try:
        comment_svg = page.locator('svg[aria-label="Comment"]')
        if comment_svg.count() > 0:
            # We know there are comments, try to find the count
            # Try parent elements for comment count
            parent = comment_svg.first.locator("xpath=ancestor::div[contains(@class, 'x')]//span").first
            if parent.count() > 0:
                text = parent.inner_text()
                value = parse_number_token(text)
                if value and value > 0:
                    evidence.value = value
                    evidence.source = "dom_related"
                    evidence.status = "OK"
                    evidence.extraction_method = "dom_ancestor_span"
                    return value, evidence
    except Exception:
        pass

    # Not found - comments might be 0 or hidden
    evidence.status = "NULL"
    evidence.null_reason = "Comment count not available (may be 0 or requires login)"
    return None, evidence
@dataclass
class AccountRow:
    no: int
    nama_kanwil: str
    url_akun: str
    manual_judul: str
    manual_link: str
    manual_reach: str
    agenda_no: str
    agenda_topic: str

@dataclass
class ScrapeRow:
    nama_kanwil: str
    url_akun: str
    post_url: str
    shortcode: str
    tanggal_postingan: Optional[datetime]
    media_type: str
    caption: str
    like_count: Optional[int]
    comment_count: Optional[int]
    total_engagement: Optional[int]
    status_periode: str
    status_scraping: str
    catatan: str
    # Extended metrics (optional)
    view_count: Optional[int] = None
    play_count: Optional[int] = None
    share_count: Optional[int] = None
    save_count: Optional[int] = None
    reach_count: Optional[int] = None

    def get_extended_display(self, field_name: str) -> str:
        """Get display value for extended field ('-' if None)."""
        value_map = {
            "view_count": self.view_count,
            "play_count": self.play_count,
            "share_count": self.share_count,
            "save_count": self.save_count,
            "reach_count": self.reach_count,
        }
        value = value_map.get(field_name)
        return str(value) if value is not None else "-"

def safe_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()

def normalize_url(url: str) -> str:
    url = safe_text(url)
    if not url:
        return ""
    if not url.startswith("http"):
        url = "https://" + url
    return url.strip()

def extract_shortcode(url: str) -> str:
    url = normalize_url(url)
    if not url:
        return ""
    try:
        parsed = urlparse(url)
        parts = [p for p in parsed.path.split("/") if p]
        if len(parts) >= 2 and parts[0] in ["p", "reel", "tv"]:
            return parts[1]
        if len(parts) >= 3 and parts[1] in ["p", "reel", "tv"]:
            return parts[2]
        return ""
    except Exception:
        return ""

def detect_media_type(post_url: str, page_content: str = None) -> str:
    """
    Deteksi media type dari URL dan optional page content.
    Returns distinct values: image, video, carousel, reels, unknown
    """
    url_lower = post_url.lower()

    # Priority 1: URL-based detection (most reliable)
    if "/reel/" in url_lower:
        return "reels"
    if "/tv/" in url_lower:
        return "video"
    if "/p/" in url_lower:
        # Check page content untuk carousel indicator
        if page_content:
            content_lower = page_content.lower()
            # Carousel indicators
            carousel_indicators = [
                "multiple",
                "carousel",
                "album",
                "slideshow",
                "swipe",
                "view carousel",
                "photo"
            ]
            # Jika ada indicator carousel dan multiple images
            if any(ind in content_lower for ind in carousel_indicators):
                return "carousel"
        return "image"

    return "unknown"


def detect_media_type_from_url(post_url: str) -> str:
    """
    Alias untuk backward compatibility.
    Deteksi media type HANYA dari URL.
    """
    return detect_media_type(post_url)

def clean_caption(text: str, max_len: int = 800) -> str:
    text = safe_text(text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return ""
    return text[:max_len]

def parse_dt_to_wib(raw: str) -> Optional[datetime]:
    """
    Parse datetime string ke datetime object (naive, tanpa timezone).
    Excel tidak mendukung timezone-aware datetime.
    """
    raw = safe_text(raw)
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        # Always strip timezone to make it naive for Excel compatibility
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt
    except Exception:
        return None

def parse_number_token(token: str) -> Optional[int]:
    token = safe_text(token).lower()
    if not token:
        return None
    token = token.replace(" ", "")
    multiplier = 1
    suffix_map = {
        "k": 1_000,
        "rb": 1_000,
        "ribu": 1_000,
        "m": 1_000_000,
        "jt": 1_000_000,
        "juta": 1_000_000,
    }
    for suffix, value in suffix_map.items():
        if token.endswith(suffix):
            multiplier = value
            token = token[: -len(suffix)]
            break
    token = token.strip()
    if not token:
        return None
    if multiplier > 1:
        token = token.replace(",", ".")
        try:
            return int(float(token) * multiplier)
        except Exception:
            return None
    if "," in token and "." in token:
        token = token.replace(",", "")
    elif "," in token:
        before, after = token.split(",", 1)
        if len(after) == 3:
            token = before + after
        else:
            token = token.replace(",", ".")
    elif "." in token:
        before, after = token.split(".", 1)
        if len(after) == 3:
            token = before + after
    try:
        return int(float(token))
    except Exception:
        return None

def parse_engagement_from_text(text: str) -> Tuple[Optional[int], Optional[int]]:
    text = safe_text(text)
    if not text:
        return None, None
    normalized = text.lower()
    number_pattern = r"([\d]+(?:[.,][\d]+)?(?:\s*(?:k|m|rb|ribu|jt|juta))?)"
    like_patterns = [
        rf"{number_pattern}\s+likes?",
        rf"{number_pattern}\s+suka",
    ]
    comment_patterns = [
        rf"{number_pattern}\s+comments?",
        rf"{number_pattern}\s+komentar",
    ]
    like_count = None
    comment_count = None
    for pattern in like_patterns:
        match = re.search(pattern, normalized)
        if match:
            like_count = parse_number_token(match.group(1))
            break
    for pattern in comment_patterns:
        match = re.search(pattern, normalized)
        if match:
            comment_count = parse_number_token(match.group(1))
            break
    return like_count, comment_count

def extract_caption_from_meta(*meta_values: str) -> str:
    combined = ""
    for value in meta_values:
        value = safe_text(value)
        if value:
            combined = value
            break
    if not combined:
        return ""
    text = re.sub(r"\s+", " ", combined).strip()
    if ": " in text:
        possible_caption = text.split(": ", 1)[-1]
        possible_caption = possible_caption.strip().strip('"').strip("“").strip("”")
        if possible_caption:
            return clean_caption(possible_caption)
    if " - " in text:
        possible_caption = text.split(" - ", 1)[-1]
        possible_caption = possible_caption.strip().strip('"').strip("“").strip("”")
        if possible_caption:
            return clean_caption(possible_caption)
    return clean_caption(text)

def status_periode(tanggal: Optional[datetime], period_start: datetime, period_end: datetime) -> str:
    if tanggal is None:
        return "Perlu Cek Manual"
    if period_start <= tanggal <= period_end:
        return "Masuk Periode"
    return "Di Luar Periode"


@dataclass
class FieldEvidence:
    """Evidence untuk satu field yang di-extract."""
    field_name: str
    value: Any = None
    source: Optional[str] = None
    selector_used: Optional[str] = None
    extraction_method: Optional[str] = None
    status: str = "NULL"
    null_reason: Optional[str] = None
    attempted_selectors: List[str] = field(default_factory=list)

    def is_ok(self) -> bool:
        return self.value is not None and self.status == "OK"


@dataclass
class EnhancedScrapeResult:
    """
    Enhanced result dengan field-level evidence tracking.
    Untuk debugging dan maintenance scraper.
    """
    # Basic
    nama_kanwil: str = ""
    url_akun: str = ""
    post_url: str = ""
    shortcode: str = ""

    # Field evidence
    caption_evidence: Optional[FieldEvidence] = None
    timestamp_evidence: Optional[FieldEvidence] = None
    media_type_evidence: Optional[FieldEvidence] = None
    like_evidence: Optional[FieldEvidence] = None
    comment_evidence: Optional[FieldEvidence] = None

    # Aggregated values
    caption: str = ""
    timestamp: Optional[datetime] = None
    media_type: str = "UNKNOWN"
    media_url: str = ""
    alt_text: str = ""
    like_count: Optional[int] = None
    comment_count: Optional[int] = None
    total_engagement: Optional[int] = None

    # Extended metrics (optional, best-effort)
    view_count: Optional[int] = None
    play_count: Optional[int] = None
    share_count: Optional[int] = None
    save_count: Optional[int] = None
    reach_count: Optional[int] = None

    # Extended metrics evidence
    view_evidence: Optional[FieldEvidence] = None
    play_evidence: Optional[FieldEvidence] = None
    share_evidence: Optional[FieldEvidence] = None
    save_evidence: Optional[FieldEvidence] = None
    reach_evidence: Optional[FieldEvidence] = None

    # Insight availability note
    insight_note: str = ""

    # Status
    status_scraping: str = "PENDING"
    status_periode: str = ""
    catatan: str = ""

    # Debug
    html_saved: bool = False
    screenshot_saved: bool = False

    def get_field_status(self, field_name: str) -> FieldEvidence:
        """Get evidence untuk field tertentu."""
        evidence_map = {
            "caption": self.caption_evidence,
            "timestamp": self.timestamp_evidence,
            "media_type": self.media_type_evidence,
            "like_count": self.like_evidence,
            "comment_count": self.comment_evidence,
            "view_count": self.view_evidence,
            "play_count": self.play_evidence,
            "share_count": self.share_evidence,
            "save_count": self.save_evidence,
            "reach_count": self.reach_evidence,
        }
        return evidence_map.get(field_name, FieldEvidence(field_name=field_name))

    def apply_extended_metrics(self, metrics: ExtendedMetrics):
        """
        Apply extended metrics to this result.
        Safe to call even if metrics is None.
        """
        if metrics is None:
            return

        # Apply standard metrics (override if better data found)
        if metrics.like_count is not None and (self.like_count is None or self.like_count == 0):
            self.like_count = metrics.like_count

        if metrics.comment_count is not None and (self.comment_count is None or self.comment_count == 0):
            self.comment_count = metrics.comment_count

        # Apply extended metrics (optional, don't overwrite if already set)
        if self.view_count is None and metrics.view_count is not None:
            self.view_count = metrics.view_count
            self.view_evidence = metrics.view_evidence

        if self.play_count is None and metrics.play_count is not None:
            self.play_count = metrics.play_count
            self.play_evidence = metrics.play_evidence

        if self.share_count is None and metrics.share_count is not None:
            self.share_count = metrics.share_count
            self.share_evidence = metrics.share_evidence

        if self.save_count is None and metrics.save_count is not None:
            self.save_count = metrics.save_count
            self.save_evidence = metrics.save_evidence

        if self.reach_count is None and metrics.reach_count is not None:
            self.reach_count = metrics.reach_count
            self.reach_evidence = metrics.reach_evidence

        # Set insight note if applicable
        if metrics.insight_note:
            self.insight_note = metrics.insight_note

        # Recalculate total_engagement (only if changed)
        if self.like_count is not None or self.comment_count is not None:
            new_engagement = (self.like_count or 0) + (self.comment_count or 0)
            if new_engagement > 0:
                self.total_engagement = new_engagement

    def to_scrape_row(self) -> ScrapeRow:
        """Convert ke ScrapeRow standar."""
        return ScrapeRow(
            nama_kanwil=self.nama_kanwil,
            url_akun=self.url_akun,
            post_url=self.post_url,
            shortcode=self.shortcode,
            tanggal_postingan=self.timestamp,
            media_type=self.media_type,
            caption=self.caption,
            like_count=self.like_count,
            comment_count=self.comment_count,
            total_engagement=self.total_engagement,
            status_periode=self.status_periode,
            status_scraping=self.status_scraping,
            catatan=self.catatan,
            # Extended metrics
            view_count=self.view_count,
            play_count=self.play_count,
            share_count=self.share_count,
            save_count=self.save_count,
            reach_count=self.reach_count,
        )

def load_accounts_from_excel(file_bytes: bytes) -> List[AccountRow]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)

    sheet = None

    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == SHEET_SOURCE.lower():
            sheet = workbook[sheet_name]
            break

    if sheet is None:
        for candidate_sheet in workbook.worksheets:
            found_instagram = False

            for row in candidate_sheet.iter_rows(values_only=True):
                for value in row:
                    if "instagram.com" in safe_text(value).lower():
                        found_instagram = True
                        break

                if found_instagram:
                    break

            if found_instagram:
                sheet = candidate_sheet
                break

    if sheet is None:
        raise ValueError("Tidak ada sheet yang berisi link Instagram. Pastikan file berisi URL instagram.com.")

    def to_profile_url(value: str) -> str:
        value = normalize_url(value)

        if not value:
            return ""

        try:
            parsed = urlparse(value)
            parts = [part for part in parsed.path.split("/") if part]

            if not parts:
                return ""

            if parts[0] in ["p", "reel", "tv", "stories"]:
                return ""

            username = parts[0].strip().lower()

            if not username:
                return ""

            return f"https://www.instagram.com/{username}/"

        except Exception:
            return ""

    def username_from_url(value: str) -> str:
        profile_url = to_profile_url(value)

        if not profile_url:
            return ""

        parsed = urlparse(profile_url)
        parts = [part for part in parsed.path.split("/") if part]

        if not parts:
            return ""

        return parts[0].strip()

    accounts = []
    seen_key = set()

    for row_index in range(1, sheet.max_row + 1):
        row_values = [
            safe_text(sheet.cell(row=row_index, column=col_index).value)
            for col_index in range(1, sheet.max_column + 1)
        ]

        instagram_columns = []

        for col_index, value in enumerate(row_values, start=1):
            if "instagram.com" in value.lower():
                instagram_columns.append(col_index)

        if not instagram_columns:
            continue

        for url_col in instagram_columns:
            raw_url = safe_text(sheet.cell(row=row_index, column=url_col).value)
            profile_url = to_profile_url(raw_url)

            if not profile_url:
                continue

            nama = ""

            if url_col > 1:
                nama = safe_text(sheet.cell(row=row_index, column=url_col - 1).value)

            if "kanwil" not in nama.lower() or "djpb" not in nama.lower():
                nama = safe_text(sheet.cell(row=row_index, column=2).value)

            if "kanwil" not in nama.lower() or "djpb" not in nama.lower():
                for value in row_values:
                    value_lower = value.lower()

                    if "kanwil" in value_lower and "djpb" in value_lower:
                        nama = value
                        break

            if not nama:
                username = username_from_url(profile_url)
                nama = f"Kanwil DJPb {username}"

            key = profile_url.lower()

            if key in seen_key:
                continue

            seen_key.add(key)

            accounts.append(
                AccountRow(
                    no=len(accounts) + 1,
                    nama_kanwil=nama,
                    url_akun=profile_url,
                    manual_judul=safe_text(sheet.cell(row=row_index, column=6).value),
                    manual_link=normalize_url(sheet.cell(row=row_index, column=7).value),
                    manual_reach=safe_text(sheet.cell(row=row_index, column=9).value),
                    agenda_no=safe_text(sheet.cell(row=row_index, column=10).value),
                    agenda_topic=safe_text(sheet.cell(row=row_index, column=11).value),
                )
            )
    if not accounts:
        raise ValueError("Tidak ada akun Instagram Kanwil DJPb yang terbaca. Pastikan file memiliki nama Kanwil dan URL instagram.com.")
    return accounts
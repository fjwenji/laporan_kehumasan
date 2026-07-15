"""
Export Service - Enhanced Excel export dengan 6 sheet.

Sheet yang dihasilkan:
1. SUMMARY - Statistik ringkasan
2. KANWIL - Data postingan Kanwil
3. KPPN - Header saja (data belum dipakai)
4. ALL_POSTS - Semua postingan
5. COVERAGE - Status per akun
6. DATA_QUALITY - Quality check
"""

from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.export_helpers import (
    format_null_display,
    format_number,
    compute_engagement,
    truncate_caption,
    clean_text_for_excel,
    safe_datetime,
    format_date_display,
    format_date_only,
    get_media_type_display,
    determine_coverage_status,
    is_valid_instagram_url,
    calculate_summary_stats,
    calculate_data_quality,
)
from src.database import get_db_cursor
from src.notification_service import send_telegram_message


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
SUCCESS_FILL = PatternFill("solid", fgColor="C6EFCE")
WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def style_header_row(sheet, row: int, num_cols: int):
    """Style header row."""
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_row(sheet, row: int, num_cols: int):
    """Style data row."""
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def apply_status_color(cell, status: str):
    """Apply color based on status."""
    if not status:
        return
    status_lower = status.lower()
    if "success" in status_lower and "partial" not in status_lower:
        cell.fill = SUCCESS_FILL
    elif "partial" in status_lower or "null" in status_lower or "review" in status_lower:
        cell.fill = WARNING_FILL
    elif any(x in status_lower for x in ["fail", "error", "login", "not_found", "rate", "timeout"]):
        cell.fill = ERROR_FILL


def set_column_widths(sheet, widths: Dict[str, int]):
    """Set column widths."""
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width


def get_kanwil_accounts(limit: int = 34) -> List[Dict]:
    """Get Kanwil accounts from database."""
    with get_db_cursor(commit=False) as cursor:
        cursor.execute("""
            SELECT id, username, nama_unit, kategori_unit, wilayah,
                   is_active, last_checked_at, last_latest_scrape_at
            FROM accounts
            WHERE kategori_unit = 'KANWIL'
              AND username LIKE 'djpb%%'
              AND is_active = 1
            ORDER BY nama_unit
            LIMIT %s
        """, (limit,))
        return cursor.fetchall()


def get_posts_for_period(
    period_start: date,
    period_end: date,
    scope: str = "kanwil"
) -> List[Dict]:
    """
    Get posts for export period.

    Args:
        period_start: Start date
        period_end: End date
        scope: 'kanwil', 'kppn', atau 'all'

    Returns:
        List of post dictionaries
    """
    scope_filter = ""
    if scope == "kanwil":
        scope_filter = "AND a.kategori_unit = 'KANWIL' AND a.username LIKE 'djpb%%'"
    elif scope == "kppn":
        scope_filter = "AND a.kategori_unit = 'KPPN'"

    with get_db_cursor(commit=False) as cursor:
        cursor.execute(f"""
            SELECT p.*,
                   a.nama_unit as account_nama_unit,
                   a.username as account_username,
                   a.kategori_unit,
                   a.wilayah,
                   COALESCE(p.media_type_normalized, 'unknown') AS media_type_display
            FROM posts p
            LEFT JOIN accounts a ON p.username = a.username
            WHERE DATE(p.timestamp) BETWEEN %s AND %s
              {scope_filter}
            ORDER BY p.timestamp DESC
        """, (period_start, period_end))
        return cursor.fetchall()


def get_posts_by_scope(
    period_start: date,
    period_end: date,
    scope: str = "kanwil"
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    Get posts split by scope (Kanwil, KPPN, All).

    Returns: (kanwil_posts, kppn_posts, all_posts)
    """
    all_posts = get_posts_for_period(period_start, period_end, "all")

    kanwil_posts = []
    kppn_posts = []

    for post in all_posts:
        kategori = post.get('kategori_unit', '').upper()
        username = (post.get('username') or '').lower()

        if kategori == 'KANWIL' or 'djpb' in username:
            kanwil_posts.append(post)
        elif kategori == 'KPPN':
            kppn_posts.append(post)

    return kanwil_posts, kppn_posts, all_posts


def get_account_post_counts(
    accounts: List[Dict],
    period_start: date,
    period_end: date
) -> Dict[str, Dict]:
    """
    Get post counts per account for coverage calculation.

    Returns:
        Dict[username, {'count', 'first', 'last', 'dates'}]
    """
    if not accounts:
        return {}

    usernames = [acc['username'] for acc in accounts]
    placeholders = ','.join(['%s'] * len(usernames))

    with get_db_cursor(commit=False) as cursor:
        cursor.execute(f"""
            SELECT username,
                   COUNT(*) as post_count,
                   MIN(timestamp) as first_post,
                   MAX(timestamp) as last_post
            FROM posts
            WHERE username IN ({placeholders})
              AND DATE(timestamp) BETWEEN %s AND %s
            GROUP BY username
        """, usernames + [period_start, period_end])

        result = {}
        for row in cursor.fetchall():
            result[row['username']] = {
                'count': row['post_count'],
                'first': row['first_post'],
                'last': row['last_post']
            }
        return result


def build_summary_sheet(
    wb: Workbook,
    stats: Dict[str, Any],
    scope: str,
    period_start: date,
    period_end: date,
    total_accounts: int
):
    """Build SUMMARY sheet."""
    ws = wb.active
    ws.title = "SUMMARY"

    row = 1

    # Title
    ws.cell(row=row, column=1, value="LAPORAN MONITORING INSTAGRAM DJPb")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 1

    ws.cell(row=row, column=1, value=f"Periode: {period_start.strftime('%d/%m/%Y')} - {period_end.strftime('%d/%m/%Y')}")
    row += 1

    ws.cell(row=row, column=1, value=f"Scope: {scope.upper()}")
    row += 1

    ws.cell(row=row, column=1, value=f"Tanggal Export: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    row += 2

    # Section: Statistics
    ws.cell(row=row, column=1, value="STATISTIK")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    stats_rows = [
        ("Total Akun dalam Export", total_accounts),
        ("Total Postingan", stats.get('total_posts', 0)),
        ("", ""),
        ("Breakdown Media Type:", ""),
        ("  - Image", stats.get('image_count', 0)),
        ("  - Carousel", stats.get('carousel_count', 0)),
        ("  - Reels", stats.get('reels_count', 0)),
        ("  - Video", stats.get('video_count', 0)),
        ("  - Unknown", stats.get('unknown_count', 0)),
        ("", ""),
        ("Engagement:", ""),
        ("  - Total Like", format_number(stats.get('total_likes', 0))),
        ("  - Total Komentar", format_number(stats.get('total_comments', 0))),
        ("  - Total Engagement", format_number(stats.get('total_engagement', 0))),
        ("", ""),
        ("Views (Reels/Video):", ""),
        ("  - Total Views", format_number(stats.get('total_views', 0))),
        ("  - Post dengan View", stats.get('posts_with_views', 0)),
        ("  - Post tanpa View", stats.get('posts_without_views', 0)),
        ("", ""),
        ("Catatan:", "view_count bersifat nullable - tidak semua post memiliki data views"),
    ]

    for label, value in stats_rows:
        ws.cell(row=row, column=1, value=label)
        if value:
            ws.cell(row=row, column=2, value=value)
        row += 1

    set_column_widths(ws, {"A": 35, "B": 25})


def build_posts_sheet(
    wb: Workbook,
    sheet_name: str,
    posts: List[Dict],
    include_view_count: bool = True
):
    """Build KANWIL, KPPN, atau ALL_POSTS sheet."""
    ws = wb.create_sheet(sheet_name)

    # Headers - 19 columns
    headers = [
        "No.", "Username", "Nama Unit", "Jenis Unit",
        "Post URL", "Shortcode",
        "Caption", "Caption Ringkas",
        "Posted At", "Media Type",
        "Like", "Comment", "Views", "Play",
        "Engagement",
        "Is New Post", "Status Scraping", "Null Reason",
        "Last Scraped At"
    ]

    num_cols = len(headers)

    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, num_cols)

    # Write data
    row_num = 2
    for idx, post in enumerate(posts, 1):
        username = post.get('username', '-') or '-'
        nama_unit = post.get('nama_unit') or post.get('account_nama_unit', '-') or '-'
        kategori = post.get('kategori_unit', '-') or '-'
        post_url = post.get('post_url', '-') or '-'
        shortcode = post.get('shortcode', '-') or '-'
        caption = clean_text_for_excel(post.get('caption', ''))
        caption_ringkas = truncate_caption(caption, 100)
        posted_at = safe_datetime(post.get('timestamp'))
        media_type = get_media_type_display(
            post.get('media_type'),
            post.get('media_type_display')
        )
        like_count = post.get('like_count')
        comment_count = post.get('comments_count')
        view_count = post.get('view_count')
        play_count = post.get('play_count')
        engagement = compute_engagement(like_count, comment_count)
        is_new = "Ya" if post.get('is_new_post') else "Tidak"
        status = post.get('status_scraping', '-') or '-'
        null_reason = post.get('null_reason', '') or ''
        last_scraped = format_date_display(post.get('last_scraped_at'))

        # Handle view_count NULL -> display as "-"
        view_display = format_null_display(view_count)

        row_data = [
            idx,
            f"@{username}" if username != '-' else '-',
            nama_unit[:50],
            kategori,
            post_url if post_url != '-' else '-',
            shortcode,
            caption,
            caption_ringkas,
            posted_at,
            media_type,
            format_null_display(like_count, return_zero=True),
            format_null_display(comment_count, return_zero=True),
            view_display,  # NULL -> "-"
            format_null_display(play_count),
            format_null_display(engagement, return_zero=True),
            is_new,
            status,
            null_reason[:100],
            last_scraped,
        ]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

            # Apply status color to Status Scraping column (17)
            if col == 17:
                apply_status_color(ws.cell(row=row_num, column=col), value)

        style_data_row(ws, row_num, num_cols)
        row_num += 1

    # Column widths
    widths = {
        "A": 6, "B": 18, "C": 35, "D": 12,
        "E": 60, "F": 20,
        "G": 70, "H": 35,
        "I": 18, "J": 14,
        "K": 10, "L": 10, "M": 12, "N": 10,
        "O": 12,
        "P": 10, "Q": 20, "R": 30,
        "S": 20,
    }
    set_column_widths(ws, widths)
    ws.freeze_panes = "A2"


def build_coverage_sheet(
    wb: Workbook,
    accounts: List[Dict],
    post_counts: Dict[str, Dict],
    period_start: date,
    period_end: date
):
    """Build COVERAGE sheet."""
    ws = wb.create_sheet("COVERAGE")

    # Headers
    headers = [
        "No.", "Username", "Nama Unit", "Jenis Unit", "Wilayah",
        "Total Post", "First Post", "Last Post",
        "Status", "Last Scrape", "Note"
    ]
    num_cols = len(headers)

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, num_cols)

    row_num = 2
    for idx, acc in enumerate(accounts, 1):
        username = acc.get('username', '-') or '-'
        nama_unit = acc.get('nama_unit', '-') or '-'
        kategori = acc.get('kategori_unit', '-') or '-'
        wilayah = acc.get('wilayah', '-') or '-'
        last_scrape = format_date_display(acc.get('last_latest_scrape_at'))

        post_info = post_counts.get(username, {'count': 0, 'first': None, 'last': None})
        post_count = post_info['count']
        first_post = post_info['first']
        last_post = post_info['last']

        has_data = post_count > 0
        coverage_status, note = determine_coverage_status(
            has_data, first_post, last_post,
            period_start, period_end
        )

        row_data = [
            idx,
            f"@{username}",
            nama_unit[:50],
            kategori,
            wilayah[:20],
            post_count,
            format_date_only(first_post) if first_post else "-",
            format_date_only(last_post) if last_post else "-",
            coverage_status,
            last_scrape,
            note,
        ]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

            # Color status column (8)
            if col == 8:
                if coverage_status == "COMPLETE":
                    ws.cell(row=row_num, column=col).fill = SUCCESS_FILL
                elif coverage_status == "PARTIAL":
                    ws.cell(row=row_num, column=col).fill = WARNING_FILL
                else:
                    ws.cell(row=row_num, column=col).fill = ERROR_FILL

        style_data_row(ws, row_num, num_cols)
        row_num += 1

    # Column widths
    widths = {
        "A": 5, "B": 18, "C": 35, "D": 12, "E": 20,
        "F": 12, "G": 15, "H": 15,
        "I": 12, "J": 18, "K": 40,
    }
    set_column_widths(ws, widths)
    ws.freeze_panes = "A2"


def build_data_quality_sheet(
    wb: Workbook,
    quality: Dict[str, Any]
):
    """Build DATA_QUALITY sheet."""
    ws = wb.create_sheet("DATA_QUALITY")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="DATA QUALITY CHECK")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 2

    # Missing fields
    ws.cell(row=row, column=1, value="MISSING DATA:")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    missing_rows = [
        ("Post tanpa timestamp", quality.get('missing_timestamp', 0)),
        ("Post tanpa caption", quality.get('missing_caption', 0)),
        ("Post tanpa like_count", quality.get('missing_likes', 0)),
        ("Post tanpa comments_count", quality.get('missing_comments', 0)),
    ]

    for label, count in missing_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1

    # Media type issues
    ws.cell(row=row, column=1, value="MEDIA TYPE ISSUES:")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    media_rows = [
        ("Reels tanpa view_count", quality.get('reels_without_views', 0)),
        ("Media type Unknown", quality.get('unknown_media_type', 0)),
    ]

    for label, count in media_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1

    # URL issues
    ws.cell(row=row, column=1, value="URL/DATA ISSUES:")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    url_rows = [
        ("Invalid URL (chrome-error)", quality.get('invalid_urls', 0)),
        ("Duplicate shortcode", quality.get('duplicate_shortcodes', 0)),
    ]

    for label, count in url_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 2

    # Recommendations
    ws.cell(row=row, column=1, value="REKOMENDASI:")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    recommendations = quality.get('recommendations', [])
    if recommendations:
        for rec in recommendations:
            ws.cell(row=row, column=1, value=f"• {rec}")
            row += 1
    else:
        ws.cell(row=row, column=1, value="✅ Tidak ada masalah critical terdeteksi")
        row += 1

    set_column_widths(ws, {"A": 50, "B": 15})


def create_export(
    period_start: date,
    period_end: date,
    scope: str = "kanwil",
    account_limit: int = 34,
    include_kppn: bool = True,
    include_all: bool = True,
) -> bytes:
    """
    Create enhanced Excel export dengan 6 sheet.

    Args:
        period_start: Start date
        period_end: End date
        scope: 'kanwil', 'kppn', atau 'all'
        account_limit: Max accounts to include
        include_kppn: Include KPPN sheet (even if empty)
        include_all: Include ALL_POSTS sheet

    Returns:
        Excel file as bytes
    """
    wb = Workbook()

    # Get data
    kanwil_posts, kppn_posts, all_posts = get_posts_by_scope(period_start, period_end, scope)
    kanwil_accounts = get_kanwil_accounts(limit=account_limit)
    post_counts = get_account_post_counts(kanwil_accounts, period_start, period_end)

    # Calculate stats
    summary_stats = calculate_summary_stats(all_posts if scope == "all" else kanwil_posts)

    # Build sheets

    # 1. SUMMARY
    build_summary_sheet(
        wb, summary_stats, scope,
        period_start, period_end,
        len(kanwil_accounts)
    )

    # 2. KANWIL (always include)
    build_posts_sheet(wb, "KANWIL", kanwil_posts)

    # 3. KPPN (include header even if empty)
    if include_kppn:
        build_posts_sheet(wb, "KPPN", kppn_posts)
    else:
        ws = wb.create_sheet("KPPN")
        ws.cell(row=1, column=1, value="Data KPPN belum dipakai dalam export ini")

    # 4. ALL_POSTS
    if include_all:
        build_posts_sheet(wb, "ALL_POSTS", all_posts)

    # 5. COVERAGE
    build_coverage_sheet(wb, kanwil_accounts, post_counts, period_start, period_end)

    # 6. DATA_QUALITY
    quality = calculate_data_quality(all_posts if scope == "all" else kanwil_posts)
    build_data_quality_sheet(wb, quality)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_export_filename(period_start: date, period_end: date, scope: str = "kanwil") -> str:
    """Generate export filename."""
    scope_prefix = {
        "kanwil": "Kanwil",
        "kppn": "KPPN",
        "all": "All",
    }.get(scope, scope.title())

    return f"Mayz_Export_{scope_prefix}_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}.xlsx"


def notify_export_complete(
    filename: str,
    scope: str,
    period_start: date,
    period_end: date,
    total_posts: int,
    total_accounts: int
) -> Tuple[bool, str]:
    """
    Kirim notifikasi Telegram setelah export selesai.

    Returns:
        Tuple[bool, str]: (success, message)
    """
    from src.database import get_setting

    # Check if Telegram enabled
    telegram_enabled = get_setting("TELEGRAM_ENABLED", "false").lower() == "true"
    if not telegram_enabled:
        return False, "Telegram tidak diaktifkan"

    scope_display = {
        "kanwil": "34 Kanwil",
        "kppn": "KPPN",
        "all": "Semua Akun",
    }.get(scope, scope.title())

    message = (
        f"📊 <b>Export Excel Selesai</b>\n\n"
        f"<b>File:</b> {filename}\n"
        f"<b>Scope:</b> {scope_display}\n"
        f"<b>Periode:</b> {period_start.strftime('%d/%m/%Y')} s/d {period_end.strftime('%d/%m/%Y')}\n"
        f"<b>Total Akun:</b> {total_accounts}\n"
        f"<b>Total Postingan:</b> {total_posts:,}\n\n"
        f"<i>Waktu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>"
    )

    return send_telegram_message(message)

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from src.config import BASE_HEADERS, EXTRA_FIELD_MAP, SHEET_OUTPUT, SHEET_RAW
from src.parser import AccountRow, ScrapeRow

HEADER_ROW = 6
TITLE_ROW = 8
DATA_START_ROW = 9
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, size=11)
BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)

def style_header(sheet, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_body(sheet, min_row: int, max_row: int, min_col: int, max_col: int):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

def set_widths(sheet, widths: Dict[str, int]):
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

def build_output(accounts: List[AccountRow], rows: List[ScrapeRow], selected_fields: List[str], only_period: bool, include_raw: bool) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_OUTPUT
    headers = BASE_HEADERS + selected_fields
    last_col_letter = chr(64 + min(len(headers), 26))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"] = "DAFTAR LINK PELAPORAN EKSIS VERTIKAL"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet["A2"] = "DIREKTORAT JENDERAL PERBENDAHARAAN"
    sheet["A2"].font = Font(bold=True, size=12)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet["A4"] = "Juni 2026"
    sheet["A4"].font = Font(bold=True)
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=HEADER_ROW, column=index).value = header
    style_header(sheet, HEADER_ROW, 1, len(headers))
    sheet.cell(row=TITLE_ROW, column=1).value = "Hasil Scraping Public Instagram"
    sheet.cell(row=TITLE_ROW, column=1).font = SECTION_FONT

    def safe_datetime(dt: Optional[datetime]) -> Optional[datetime]:
        """Ensure datetime is naive (no timezone) for Excel compatibility."""
        if dt is None:
            return None
        if dt.tzinfo is not None:
            return dt.replace(tzinfo=None)
        return dt

    grouped = {}
    for item in rows:
        grouped.setdefault(item.nama_kanwil, []).append(item)
    output_row = DATA_START_ROW
    number = 1
    for account in accounts:
        candidates = [item for item in grouped.get(account.nama_kanwil, []) if item.post_url]
        if only_period:
            candidates = [item for item in candidates if item.status_periode == "Masuk Periode"]
        for item in candidates:
            values = [
                number,
                account.nama_kanwil,
                account.url_akun,
                safe_datetime(item.tanggal_postingan),
                "Medsos",
                item.caption or account.manual_judul or "Perlu Cek Manual",
                item.post_url,
                "IG",
                "",
                account.agenda_no or "",
                account.agenda_topic or "",
            ]
            for field in selected_fields:
                attr = EXTRA_FIELD_MAP.get(field)
                value = getattr(item, attr, "") if attr else ""
                values.append(value if value is not None else "")
            for col, value in enumerate(values, start=1):
                sheet.cell(row=output_row, column=col).value = value
            if item.post_url:
                sheet.cell(row=output_row, column=7).hyperlink = item.post_url
                sheet.cell(row=output_row, column=7).style = "Hyperlink"
            number += 1
            output_row += 1
    if output_row == DATA_START_ROW:
        sheet.cell(row=DATA_START_ROW, column=1).value = "Tidak ada data postingan yang berhasil dipetakan."
    style_body(sheet, DATA_START_ROW, max(output_row - 1, DATA_START_ROW), 1, len(headers))
    for row in range(DATA_START_ROW, max(output_row, DATA_START_ROW + 1)):
        sheet.cell(row=row, column=4).number_format = "dd mmmm yyyy hh:mm"
    set_widths(sheet, {
        "A": 8,
        "B": 40,
        "C": 50,
        "D": 22,
        "E": 16,
        "F": 80,
        "G": 65,
        "H": 18,
        "I": 22,
        "J": 36,
        "K": 44,
        "L": 16,
        "M": 16,
        "N": 18,
        "O": 24,
        "P": 24,
        "Q": 24,
        "R": 24,
        "S": 60,
    })
    sheet.freeze_panes = "A9"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{max(output_row - 1, DATA_START_ROW)}"
    if include_raw:
        raw = workbook.create_sheet(SHEET_RAW)
        raw_headers = [
            "nama_kanwil",
            "url_akun",
            "post_url",
            "shortcode",
            "tanggal_postingan",
            "media_type",
            "caption",
            "like_count",
            "comment_count",
            "total_engagement",
            "status_periode",
            "status_scraping",
            "catatan",
        ]
        raw.append(raw_headers)
        style_header(raw, 1, 1, len(raw_headers))
        for item in rows:
            raw.append([
                item.nama_kanwil,
                item.url_akun,
                item.post_url,
                item.shortcode,
                safe_datetime(item.tanggal_postingan),
                item.media_type,
                item.caption,
                item.like_count if item.like_count is not None else "",
                item.comment_count if item.comment_count is not None else "",
                item.total_engagement if item.total_engagement is not None else "",
                item.status_periode,
                item.status_scraping,
                item.catatan,
            ])
        style_body(raw, 2, max(raw.max_row, 2), 1, len(raw_headers))
        for row in range(2, raw.max_row + 1):
            raw.cell(row=row, column=5).number_format = "dd mmmm yyyy hh:mm"
        set_widths(raw, {
            "A": 40,
            "B": 50,
            "C": 65,
            "D": 22,
            "E": 22,
            "F": 24,
            "G": 80,
            "H": 14,
            "I": 16,
            "J": 18,
            "K": 22,
            "L": 22,
            "M": 60,
        })
        raw.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def save_output(content: bytes, export_dir: Path) -> Path:
    export_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"Pelaporan_Juni_2026_output_{timestamp}.xlsx"
    path.write_bytes(content)
    return path
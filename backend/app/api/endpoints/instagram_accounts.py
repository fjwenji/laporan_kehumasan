"""
Instagram Account Management API Endpoints
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from typing import Optional, List
from datetime import datetime
import re
import io

from app.api.deps import get_current_admin_user
from app.database import get_db_cursor
from app.schemas.instagram_accounts import (
    InstagramAccount, InstagramAccountCreate, InstagramAccountUpdate,
    InstagramAccountResponse, InstagramAccountFilters,
    ImportPreviewRow, ImportPreviewResponse, ImportConfirmRequest, ImportConfirmResponse,
    UsernameValidationRequest, UsernameValidationResponse,
    JenisAkun, AccountStatus
)

router = APIRouter(prefix="/api/admin/instagram-accounts", tags=["Instagram Accounts"])


# HELPER FUNCTIONS
def normalize_username(username: str) -> str:
    """Normalize Instagram username from link or plain username."""
    username = (username or "").strip()

    # If it's a URL, extract the username
    if "instagram.com" in username.lower():
        # Extract username from URL like https://www.instagram.com/djpbkemenkeu/
        import re
        match = re.search(r'instagram\.com/([^/?]+)', username)
        if match:
            username = match.group(1)

    # Remove @ if present
    username = username.lstrip("@").strip().lower()

    return username


def is_instagram_url(text: str) -> bool:
    """Check if text is an Instagram URL."""
    if not text:
        return False
    return "instagram.com" in text.lower()


def validate_username(username: str) -> tuple[bool, str]:
    """Validate Instagram username format."""
    username = normalize_username(username)

    if not username:
        return False, "Username tidak boleh kosong."

    if len(username) > 30:
        return False, "Username maksimal 30 karakter."

    if not re.match(r"^[a-zA-Z0-9._]+$", username):
        return False, "Username hanya boleh huruf, angka, titik, dan underscore."

    if username.startswith(".") or username.endswith("."):
        return False, "Username tidak boleh diawali atau diakhiri titik."

    if ".." in username:
        return False, "Username tidak boleh mengandung titik berurutan."

    return True, "Username valid."


def map_jenis_akun(jenis: str) -> str:
    """Map jenis_akun string to database value."""
    mapping = {
        "kanwil": "KANWIL",
        "kppn": "KPPN",
        "pusat": "PUSAT",
        "kanver_lainnya": "KANVER_LAINNYA",
        "kanwil ": "KANWIL",
        "kppn ": "KPPN",
        "pusat ": "PUSAT",
    }
    return mapping.get(jenis.lower().strip(), "KANWIL")


def map_status(status: str) -> bool:
    """Map status string to boolean."""
    status_lower = (status or "").lower().strip()
    return status_lower in ("aktif", "active", "1", "true", "enabled")


# CRUD ENDPOINTS

@router.get("", response_model=InstagramAccountResponse)
async def list_accounts(
    search: Optional[str] = Query(None, description="Search username or nama_unit"),
    jenis_akun: Optional[str] = Query(None, description="Filter by jenis akun"),
    status: Optional[str] = Query(None, description="Filter by status (aktif/nonaktif)"),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Get list of Instagram accounts with filtering.
    """
    with get_db_cursor(commit=False) as cursor:
        # Build query
        query = "SELECT * FROM accounts WHERE 1=1"
        count_query = "SELECT COUNT(*) as cnt FROM accounts WHERE 1=1"
        params = []
        count_params = []

        if search:
            query += " AND (username LIKE %s OR nama_unit LIKE %s)"
            count_query += " AND (username LIKE %s OR nama_unit LIKE %s)"
            search_pattern = f"%{search}%"
            params.extend([search_pattern, search_pattern])
            count_params.extend([search_pattern, search_pattern])

        if jenis_akun:
            query += " AND kategori_unit = %s"
            count_query += " AND kategori_unit = %s"
            mapped_jenis = map_jenis_akun(jenis_akun)
            params.append(mapped_jenis)
            count_params.append(mapped_jenis)

        if status is not None:
            if status.lower() == "aktif":
                query += " AND is_active = TRUE"
                count_query += " AND is_active = TRUE"
            elif status.lower() == "nonaktif":
                query += " AND is_active = FALSE"
                count_query += " AND is_active = FALSE"

        # Get total count
        cursor.execute(count_query, count_params)
        total = cursor.fetchone()["cnt"]

        # Get active/inactive counts
        cursor.execute("SELECT COUNT(*) as cnt FROM accounts WHERE is_active = TRUE")
        active_count = cursor.fetchone()["cnt"]

        cursor.execute("SELECT COUNT(*) as cnt FROM accounts WHERE is_active = FALSE")
        inactive_count = cursor.fetchone()["cnt"]

        # Get accounts
        query += " ORDER BY nama_unit ASC LIMIT %s OFFSET %s"
        params.extend([limit, offset])
        cursor.execute(query, params)
        accounts = cursor.fetchall()

    result_accounts = []
    for acc in accounts:
        result_accounts.append(InstagramAccount(
            id=acc["id"],
            username=acc["username"],
            nama_unit=acc["nama_unit"],
            jenis_akun=acc.get("kategori_unit", "KANWIL").lower(),
            status="aktif" if acc.get("is_active", True) else "nonaktif",
            notes=acc.get("notes"),
            is_active=acc.get("is_active", True),
            created_at=acc.get("created_at"),
            updated_at=acc.get("updated_at")
        ))

    return InstagramAccountResponse(
        accounts=result_accounts,
        total=total,
        active_count=active_count,
        inactive_count=inactive_count
    )


@router.get("/{account_id}", response_model=InstagramAccount)
async def get_account(
    account_id: int,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Get a single Instagram account by ID.
    """
    with get_db_cursor(commit=False) as cursor:
        cursor.execute("SELECT * FROM accounts WHERE id = %s", (account_id,))
        acc = cursor.fetchone()

    if not acc:
        raise HTTPException(status_code=404, detail="Akun tidak ditemukan.")

    return InstagramAccount(
        id=acc["id"],
        username=acc["username"],
        nama_unit=acc["nama_unit"],
        jenis_akun=acc.get("kategori_unit", "KANWIL").lower(),
        status="aktif" if acc.get("is_active", True) else "nonaktif",
        notes=acc.get("notes"),
        is_active=acc.get("is_active", True),
        created_at=acc.get("created_at"),
        updated_at=acc.get("updated_at")
    )


@router.post("", response_model=InstagramAccount)
async def create_account(
    request: InstagramAccountCreate,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Create a new Instagram account.
    """
    username = normalize_username(request.username)

    # Validate username
    valid, msg = validate_username(username)
    if not valid:
        raise HTTPException(status_code=400, detail=msg)

    # Check for duplicate
    with get_db_cursor(commit=False) as cursor:
        cursor.execute("SELECT id FROM accounts WHERE username = %s AND is_active = TRUE", (username,))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail=f"Username @{username} sudah terdaftar.")

    try:
        with get_db_cursor() as cursor:
            profile_url = f"https://www.instagram.com/{username}/"
            kategori = map_jenis_akun(request.jenis_akun.value)
            is_active = map_status(request.status.value)

            cursor.execute("""
                INSERT INTO accounts (username, nama_unit, profile_url, kategori_unit, is_active, notes)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (username, request.nama_unit, profile_url, kategori, is_active, request.notes))
            account_id = cursor.lastrowid

        with get_db_cursor(commit=False) as cursor:
            cursor.execute("SELECT * FROM accounts WHERE id = %s", (account_id,))
            acc = cursor.fetchone()

        return InstagramAccount(
            id=acc["id"],
            username=acc["username"],
            nama_unit=acc["nama_unit"],
            jenis_akun=acc.get("kategori_unit", "KANWIL").lower(),
            status="aktif" if acc.get("is_active", True) else "nonaktif",
            notes=acc.get("notes"),
            is_active=acc.get("is_active", True),
            created_at=acc.get("created_at"),
            updated_at=acc.get("updated_at")
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{account_id}", response_model=InstagramAccount)
async def update_account(
    account_id: int,
    request: InstagramAccountUpdate,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Update an existing Instagram account.
    """
    # Check account exists
    with get_db_cursor(commit=False) as cursor:
        cursor.execute("SELECT * FROM accounts WHERE id = %s", (account_id,))
        acc = cursor.fetchone()

    if not acc:
        raise HTTPException(status_code=404, detail="Akun tidak ditemukan.")

    # Validate username if changing
    if request.username:
        username = normalize_username(request.username)
        valid, msg = validate_username(username)
        if not valid:
            raise HTTPException(status_code=400, detail=msg)

        # Check for duplicate (excluding current account)
        with get_db_cursor(commit=False) as cursor:
            cursor.execute("SELECT id FROM accounts WHERE username = %s AND is_active = TRUE AND id != %s",
                         (username, account_id))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"Username @{username} sudah terdaftar.")
    else:
        username = acc["username"]

    try:
        with get_db_cursor() as cursor:
            updates = []
            values = []

            if request.username:
                updates.append("username = %s")
                values.append(normalize_username(request.username))
                updates.append("profile_url = %s")
                values.append(f"https://www.instagram.com/{normalize_username(request.username)}/")

            if request.nama_unit is not None:
                updates.append("nama_unit = %s")
                values.append(request.nama_unit)

            if request.jenis_akun is not None:
                updates.append("kategori_unit = %s")
                values.append(map_jenis_akun(request.jenis_akun.value))

            if request.status is not None:
                updates.append("is_active = %s")
                values.append(map_status(request.status.value))

            if request.notes is not None:
                updates.append("notes = %s")
                values.append(request.notes)

            if updates:
                values.append(account_id)
                cursor.execute(f"""
                    UPDATE accounts
                    SET {', '.join(updates)}, updated_at = NOW()
                    WHERE id = %s
                """, values)

        with get_db_cursor(commit=False) as cursor:
            cursor.execute("SELECT * FROM accounts WHERE id = %s", (account_id,))
            acc = cursor.fetchone()

        return InstagramAccount(
            id=acc["id"],
            username=acc["username"],
            nama_unit=acc["nama_unit"],
            jenis_akun=acc.get("kategori_unit", "KANWIL").lower(),
            status="aktif" if acc.get("is_active", True) else "nonaktif",
            notes=acc.get("notes"),
            is_active=acc.get("is_active", True),
            created_at=acc.get("created_at"),
            updated_at=acc.get("updated_at")
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{account_id}")
async def delete_account(
    account_id: int,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Delete an Instagram account.
    """
    with get_db_cursor(commit=False) as cursor:
        cursor.execute("SELECT id FROM accounts WHERE id = %s", (account_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="Akun tidak ditemukan.")

    try:
        with get_db_cursor() as cursor:
            cursor.execute("DELETE FROM accounts WHERE id = %s", (account_id,))

        return {"success": True, "message": "Akun berhasil dihapus."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{account_id}/toggle")
async def toggle_account_status(
    account_id: int,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Toggle account active status.
    """
    with get_db_cursor(commit=False) as cursor:
        cursor.execute("SELECT id, is_active FROM accounts WHERE id = %s", (account_id,))
        acc = cursor.fetchone()

    if not acc:
        raise HTTPException(status_code=404, detail="Akun tidak ditemukan.")

    try:
        new_status = not acc["is_active"]
        with get_db_cursor() as cursor:
            cursor.execute("UPDATE accounts SET is_active = %s, updated_at = NOW() WHERE id = %s",
                         (new_status, account_id))

        status_text = "diaktifkan" if new_status else "dinonaktifkan"
        return {"success": True, "message": f"Akun berhasil {status_text}."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# USERNAME VALIDATION

@router.post("/validate-username", response_model=UsernameValidationResponse)
async def validate_username_endpoint(
    request: UsernameValidationRequest,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Check if username is valid and not duplicate.
    """
    username = normalize_username(request.username)
    valid, msg = validate_username(username)

    if not valid:
        return UsernameValidationResponse(
            valid=False,
            normalized_username=username,
            message=msg,
            is_duplicate=False
        )

    # Check for duplicate
    with get_db_cursor(commit=False) as cursor:
        cursor.execute("SELECT id, nama_unit, kategori_unit FROM accounts WHERE username = %s", (username,))
        existing = cursor.fetchone()

    if existing:
        return UsernameValidationResponse(
            valid=True,
            normalized_username=username,
            message=f"Username @{username} sudah terdaftar untuk {existing['nama_unit']}.",
            is_duplicate=True,
            existing_account={
                "id": existing["id"],
                "nama_unit": existing["nama_unit"],
                "kategori_unit": existing["kategori_unit"]
            }
        )

    return UsernameValidationResponse(
        valid=True,
        normalized_username=username,
        message="Username valid dan belum terdaftar.",
        is_duplicate=False
    )


# EXCEL IMPORT

@router.post("/import-preview", response_model=ImportPreviewResponse)
async def import_preview(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Preview Excel import before confirmation.
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus format Excel (.xlsx atau .xls)")

    try:
        # Read file content
        content = await file.read()

        # Import openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        rows_data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            # Skip empty rows
            if not row or all(cell is None for cell in row):
                continue

            # Get values
            username = str(row[0]).strip() if row[0] else ""
            nama_unit = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            jenis_akun = str(row[2]).strip().lower() if len(row) > 2 and row[2] else "kanwil"
            status = str(row[3]).strip().lower() if len(row) > 3 and row[3] else "aktif"
            notes = str(row[4]).strip() if len(row) > 4 and row[4] else None

            # Skip header row - also skip if it looks like a URL (header shouldn't be URL)
            header_keywords = ["username", "akun", "account", "link", "url", "instagram"]
            if username.lower() in header_keywords or is_instagram_url(username):
                continue

            # Validate
            errors = []
            is_valid = True
            is_duplicate = False
            existing_id = None

            # Validate username
            normalized = normalize_username(username)
            username_valid, username_msg = validate_username(normalized)
            if not username_valid:
                errors.append(username_msg)
                is_valid = False

            # Check duplicate
            if username_valid:
                with get_db_cursor(commit=False) as cursor:
                    cursor.execute("SELECT id, nama_unit FROM accounts WHERE username = %s", (normalized,))
                    existing = cursor.fetchone()
                    if existing:
                        is_duplicate = True
                        existing_id = existing["id"]

            # Validate nama_unit
            if not nama_unit:
                errors.append("Nama unit tidak boleh kosong")
                is_valid = False

            # Validate jenis_akun
            valid_jenis = ["kanwil", "kppn", "pusat", "kanver_lainnya"]
            if jenis_akun not in valid_jenis:
                errors.append(f"Jenis akun tidak valid: {jenis_akun}")
                is_valid = False

            rows_data.append(ImportPreviewRow(
                row_number=row_idx,
                username=normalized,
                nama_unit=nama_unit,
                jenis_akun=jenis_akun,
                status=status,
                notes=notes,
                is_valid=is_valid and not (is_duplicate and not errors),
                error_message="; ".join(errors) if errors else None,
                is_duplicate=is_duplicate,
                existing_id=existing_id
            ))

        # Count statistics
        total_rows = len(rows_data)
        valid_rows = sum(1 for r in rows_data if r.is_valid)
        duplicate_rows = sum(1 for r in rows_data if r.is_duplicate)
        invalid_rows = sum(1 for r in rows_data if not r.is_valid)

        return ImportPreviewResponse(
            total_rows=total_rows,
            valid_rows=valid_rows,
            duplicate_rows=duplicate_rows,
            invalid_rows=invalid_rows,
            rows=rows_data,
            can_proceed=valid_rows > 0
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal membaca file Excel: {str(e)}")


@router.post("/import-confirm", response_model=ImportConfirmResponse)
async def import_confirm(
    file: UploadFile = File(...),
    request: ImportConfirmRequest = Depends(),
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Confirm Excel import with duplicate handling.
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus format Excel (.xlsx atau .xls)")

    try:
        # Read file content
        content = await file.read()

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        imported = 0
        updated = 0
        skipped = 0
        failed = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            # Skip empty rows
            if not row or all(cell is None for cell in row):
                continue

            username = str(row[0]).strip() if row[0] else ""
            nama_unit = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            jenis_akun = str(row[2]).strip().lower() if len(row) > 2 and row[2] else "kanwil"
            status = str(row[3]).strip().lower() if len(row) > 3 and row[3] else "aktif"
            notes = str(row[4]).strip() if len(row) > 4 and row[4] else None

            # Skip header row - also skip if it looks like a URL (header shouldn't be URL)
            header_keywords = ["username", "akun", "account", "link", "url", "instagram"]
            if username.lower() in header_keywords or is_instagram_url(username):
                continue

            normalized = normalize_username(username)
            username_valid, _ = validate_username(normalized)

            if not username_valid or not nama_unit:
                failed += 1
                errors.append(f"Row {row_idx}: Data tidak valid")
                continue

            try:
                with get_db_cursor(commit=False) as cursor:
                    cursor.execute("SELECT id FROM accounts WHERE username = %s", (normalized,))
                    existing = cursor.fetchone()

                if existing:
                    if request.skip_duplicates:
                        skipped += 1
                        continue
                    else:
                        # Update existing
                        with get_db_cursor() as cursor:
                            cursor.execute("""
                                UPDATE accounts
                                SET nama_unit = %s, kategori_unit = %s, is_active = %s, notes = %s, updated_at = NOW()
                                WHERE username = %s
                            """, (nama_unit, map_jenis_akun(jenis_akun), map_status(status), notes, normalized))
                        updated += 1
                else:
                    # Insert new
                    profile_url = f"https://www.instagram.com/{normalized}/"
                    with get_db_cursor() as cursor:
                        cursor.execute("""
                            INSERT INTO accounts (username, nama_unit, profile_url, kategori_unit, is_active, notes)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (normalized, nama_unit, profile_url, map_jenis_akun(jenis_akun), map_status(status), notes))
                    imported += 1

            except Exception as e:
                failed += 1
                errors.append(f"Row {row_idx}: {str(e)}")

        return ImportConfirmResponse(
            success=failed == 0,
            imported=imported,
            updated=updated,
            skipped=skipped,
            failed=failed,
            errors=errors[:20]  # Limit error list
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal import Excel: {str(e)}")

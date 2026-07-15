"""
Export endpoints
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from app.api.deps import get_current_user
from app.database import get_db_cursor

router = APIRouter(prefix="/api/export", tags=["Export"])

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def style_header(sheet, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def safe_datetime(dt) -> Optional[datetime]:
    if dt is None:
        return None
    if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


def format_number(value) -> str:
    """Format number for display, return '-' if None or 0 for optional."""
    if value is None:
        return "-"
    if isinstance(value, (int, float)):
        if value == 0:
            return "-"
        return f"{value:,}"
    return str(value)


def normalize_media_type(raw_type: str) -> str:
    """Normalize media type."""
    if not raw_type:
        return "UNCLASSIFIED_REVIEW"
    normalized = raw_type.lower().strip()
    if normalized in ["image", "foto", "gambar", "img"]:
        return "IMAGE"
    elif normalized in ["carousel", "album", "sidecar", "graphsidecar"]:
        return "CAROUSEL"
    elif normalized in ["reels", "reel"]:
        return "REELS"
    elif normalized in ["video", "tv"]:
        return "VIDEO"
    return "UNCLASSIFIED_REVIEW"


@router.get("/excel")
async def export_excel(
    start: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end: date = Query(..., description="End date (YYYY-MM-DD)"),
    username: Optional[str] = Query(None, description="Filter by account (optional)"),
    current_user: dict = Depends(get_current_user)
):
    """
    Export data to Excel file.
    Creates multiple sheets: SUMMARY, ALL_POSTS, KANWIL, COVERAGE, DATA_QUALITY
    """
    with get_db_cursor(commit=False) as cursor:
        # Get posts data
        query = """
            SELECT
                p.id, p.username, a.nama_unit, p.post_url, p.shortcode,
                p.caption, p.timestamp,
                COALESCE(p.media_type_normalized, 'unknown') as media_type,
                p.like_count, p.comments_count, p.total_engagement,
                p.view_count, p.play_count, p.share_count, p.save_count,
                p.status_scraping, p.is_new_post
            FROM posts p
            LEFT JOIN accounts a ON p.username = a.username
            WHERE DATE(p.timestamp) BETWEEN %s AND %s
        """
        params = [start, end]

        if username:
            query += " AND p.username = %s"
            params.append(username)

        query += " ORDER BY p.timestamp DESC"
        cursor.execute(query, params)
        posts = cursor.fetchall()

        # Get accounts with posts
        accounts_query = """
            SELECT DISTINCT
                a.username, a.nama_unit, a.kategori_unit, a.wilayah
            FROM accounts a
            JOIN posts p ON a.username = p.username
            WHERE DATE(p.timestamp) BETWEEN %s AND %s
        """
        if username:
            accounts_query += " AND a.username = %s"
        accounts_query += " ORDER BY a.nama_unit"
        cursor.execute(accounts_query, params)
        accounts_with_posts = cursor.fetchall()

    # Create workbook
    wb = Workbook()

    # ========== SHEET 1: SUMMARY ==========
    ws_summary = wb.active
    ws_summary.title = "SUMMARY"

    # Title
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = "LAPORAN PUBLIKASI INSTAGRAM DJPb"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    ws_summary.merge_cells('A2:F2')
    ws_summary['A2'] = f"Periode: {start} s/d {end}"
    ws_summary['A2'].font = Font(bold=True)
    ws_summary['A2'].alignment = Alignment(horizontal='center')

    # Summary data
    summary_data = [
        ("Total Akun Aktif", len(accounts_with_posts)),
        ("Total Postingan", len(posts)),
        ("Total Like", sum(p.get('like_count') or 0 for p in posts)),
        ("Total Komentar", sum(p.get('comments_count') or 0 for p in posts)),
        ("Total Engagement", sum(p.get('total_engagement') or 0 for p in posts)),
        ("Total Views", sum(p.get('view_count') or 0 for p in posts)),
    ]

    # Media type summary
    media_counts = {}
    for p in posts:
        mt = normalize_media_type(p.get('media_type', ''))
        media_counts[mt] = media_counts.get(mt, 0) + 1

    row = 4
    for label, value in summary_data:
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=format_number(value) if isinstance(value, int) else value)
        row += 1

    row += 1
    ws_summary.cell(row=row, column=1, value="Komposisi Media:").font = Font(bold=True)
    row += 1
    for mt, cnt in media_counts.items():
        ws_summary.cell(row=row, column=1, value=mt)
        ws_summary.cell(row=row, column=2, value=cnt)
        row += 1

    # Set column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15

    # ========== SHEET 2: ALL_POSTS ==========
    ws_posts = wb.create_sheet("ALL_POSTS")

    headers = [
        "No", "Username", "Nama Unit", "Tanggal", "Media Type",
        "Caption", "Link", "Like", "Komentar", "Engagement", "View", "Status"
    ]

    for col, header in enumerate(headers, 1):
        ws_posts.cell(row=1, column=col, value=header)
    style_header(ws_posts, 1, 1, len(headers))

    for i, p in enumerate(posts, 1):
        row_data = [
            i,
            p.get('username', ''),
            p.get('nama_unit', ''),
            safe_datetime(p.get('timestamp')),
            normalize_media_type(p.get('media_type', '')),
            (p.get('caption') or '')[:200],  # Truncate caption
            p.get('post_url', ''),
            p.get('like_count'),
            p.get('comments_count'),
            p.get('total_engagement'),
            p.get('view_count'),  # Optional - show as is
            p.get('status_scraping', ''),
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_posts.cell(row=i+1, column=col, value=value)
            cell.border = BORDER
            if col == 4 and isinstance(value, datetime):
                cell.number_format = 'dd/mm/yyyy hh:mm'

    # Set column widths
    ws_posts.column_dimensions['A'].width = 5
    ws_posts.column_dimensions['B'].width = 20
    ws_posts.column_dimensions['C'].width = 30
    ws_posts.column_dimensions['D'].width = 18
    ws_posts.column_dimensions['E'].width = 12
    ws_posts.column_dimensions['F'].width = 50
    ws_posts.column_dimensions['G'].width = 60
    ws_posts.column_dimensions['H'].width = 10
    ws_posts.column_dimensions['I'].width = 10
    ws_posts.column_dimensions['J'].width = 12
    ws_posts.column_dimensions['K'].width = 12
    ws_posts.column_dimensions['L'].width = 15

    ws_posts.freeze_panes = 'A2'

    # ========== SHEET 3: KANWIL ==========
    ws_kanwil = wb.create_sheet("KANWIL")

    # Filter Kanwil accounts
    kanwil_posts = [p for p in posts if p.get('kategori_unit', '').upper() in ['KANWIL', '']]

    headers_kanwil = [
        "No", "Username", "Nama Unit", "Wilayah", "Jumlah Post",
        "Total Like", "Total Komentar", "Total Engagement"
    ]

    for col, header in enumerate(headers_kanwil, 1):
        ws_kanwil.cell(row=1, column=col, value=header)
    style_header(ws_kanwil, 1, 1, len(headers_kanwil))

    # Aggregate by account
    account_stats = {}
    for p in kanwil_posts:
        uname = p.get('username', '')
        if uname not in account_stats:
            account_stats[uname] = {
                'username': uname,
                'nama_unit': p.get('nama_unit', ''),
                'wilayah': p.get('wilayah', ''),
                'count': 0,
                'likes': 0,
                'comments': 0,
                'engagement': 0
            }
        account_stats[uname]['count'] += 1
        account_stats[uname]['likes'] += p.get('like_count') or 0
        account_stats[uname]['comments'] += p.get('comments_count') or 0
        account_stats[uname]['engagement'] += p.get('total_engagement') or 0

    for i, stats in enumerate(account_stats.values(), 1):
        row_data = [
            i,
            stats['username'],
            stats['nama_unit'],
            stats['wilayah'],
            stats['count'],
            stats['likes'],
            stats['comments'],
            stats['engagement']
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_kanwil.cell(row=i+1, column=col, value=value)
            cell.border = BORDER

    ws_kanwil.column_dimensions['A'].width = 5
    ws_kanwil.column_dimensions['B'].width = 25
    ws_kanwil.column_dimensions['C'].width = 35
    ws_kanwil.column_dimensions['D'].width = 20
    ws_kanwil.column_dimensions['E'].width = 15
    ws_kanwil.column_dimensions['F'].width = 15
    ws_kanwil.column_dimensions['G'].width = 15
    ws_kanwil.column_dimensions['H'].width = 18

    # ========== SHEET 4: COVERAGE ==========
    ws_coverage = wb.create_sheet("COVERAGE")

    headers_coverage = [
        "Username", "Nama Unit", "Total Post", "Post Pertama", "Post Terakhir",
        "Status Coverage", "Catatan"
    ]

    for col, header in enumerate(headers_coverage, 1):
        ws_coverage.cell(row=1, column=col, value=header)
    style_header(ws_coverage, 1, 1, len(headers_coverage))

    row = 2
    for uname, stats in account_stats.items():
        first_post = min((p.get('timestamp') for p in kanwil_posts if p.get('username') == uname and p.get('timestamp')), default=None)
        last_post = max((p.get('timestamp') for p in kanwil_posts if p.get('username') == uname and p.get('timestamp')), default=None)

        # Determine coverage status
        if stats['count'] >= 10:
            coverage_status = "COMPLETE"
            note = "Data cukup lengkap"
        elif stats['count'] > 0:
            coverage_status = "PARTIAL"
            note = "Data kurang dari 10 postingan"
        else:
            coverage_status = "NO_DATA"
            note = "Tidak ada data"

        row_data = [
            uname,
            stats['nama_unit'],
            stats['count'],
            safe_datetime(first_post) if first_post else "-",
            safe_datetime(last_post) if last_post else "-",
            coverage_status,
            note
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_coverage.cell(row=row, column=col, value=value)
            cell.border = BORDER
            if col in [4, 5] and isinstance(value, datetime):
                cell.number_format = 'dd/mm/yyyy'
        row += 1

    ws_coverage.column_dimensions['A'].width = 25
    ws_coverage.column_dimensions['B'].width = 35
    ws_coverage.column_dimensions['C'].width = 12
    ws_coverage.column_dimensions['D'].width = 15
    ws_coverage.column_dimensions['E'].width = 15
    ws_coverage.column_dimensions['F'].width = 15
    ws_coverage.column_dimensions['G'].width = 30

    # ========== SHEET 5: DATA_QUALITY ==========
    ws_quality = wb.create_sheet("DATA_QUALITY")

    headers_quality = [
        "Metrik", "Total", "Persentase", "Catatan"
    ]

    for col, header in enumerate(headers_quality, 1):
        ws_quality.cell(row=1, column=col, value=header)
    style_header(ws_quality, 1, 1, len(headers_quality))

    total_posts = len(posts)
    missing_posted_at = sum(1 for p in posts if not p.get('timestamp'))
    missing_like = sum(1 for p in posts if p.get('like_count') is None)
    missing_comment = sum(1 for p in posts if p.get('comments_count') is None)
    missing_media = sum(1 for p in posts if not p.get('media_type') or p.get('media_type') == 'unknown')
    unclassified = sum(1 for p in posts if normalize_media_type(p.get('media_type', '')) == 'UNCLASSIFIED_REVIEW')
    failed_items = sum(1 for p in posts if p.get('status_scraping') in ['FAILED', 'LOGIN_WALL', 'PAGE_LOAD_FAILED'])

    quality_data = [
        ("Total Postingan", total_posts, "-", "-"),
        ("Missing posted_at", missing_posted_at, f"{missing_posted_at/total_posts*100:.1f}%" if total_posts else "0%", "Perlu cek manual"),
        ("Missing like_count", missing_like, f"{missing_like/total_posts*100:.1f}%" if total_posts else "0%", "Instagram mungkin batasi display"),
        ("Missing comment_count", missing_comment, f"{missing_comment/total_posts*100:.1f}%" if total_posts else "0%", "Instagram mungkin batasi display"),
        ("Missing media_type", missing_media, f"{missing_media/total_posts*100:.1f}%" if total_posts else "0%", "Perlu deteksi ulang"),
        ("Unclassified Review", unclassified, f"{unclassified/total_posts*100:.1f}%" if total_posts else "0%", "Harus review manual"),
        ("Failed Scraping", failed_items, f"{failed_items/total_posts*100:.1f}%" if total_posts else "0%", "Error saat scraping"),
        ("Optional: view_count null", sum(1 for p in posts if p.get('view_count') is None), "-", "View count tidak selalu tersedia di web"),
        ("Optional: play_count null", sum(1 for p in posts if p.get('play_count') is None), "-", "Play count tidak selalu tersedia"),
        ("Optional: share_count null", sum(1 for p in posts if p.get('share_count') is None), "-", "Share count tidak selalu tersedia"),
        ("Optional: save_count null", sum(1 for p in posts if p.get('save_count') is None), "-", "Save count tidak selalu tersedia"),
    ]

    for i, (metric, total, pct, note) in enumerate(quality_data, 2):
        ws_quality.cell(row=i, column=1, value=metric)
        ws_quality.cell(row=i, column=2, value=total)
        ws_quality.cell(row=i, column=3, value=pct)
        ws_quality.cell(row=i, column=4, value=note)
        for col in range(1, 5):
            ws_quality.cell(row=i, column=col).border = BORDER

    ws_quality.column_dimensions['A'].width = 30
    ws_quality.column_dimensions['B'].width = 12
    ws_quality.column_dimensions['C'].width = 15
    ws_quality.column_dimensions['D'].width = 40

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Log export
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                INSERT INTO export_logs (exported_by, period_start, period_end, record_count, created_at)
                VALUES (%s, %s, %s, %s, NOW())
            """, (current_user["username"], start, end, len(posts)))
    except:
        pass  # Don't fail if logging fails

    filename = f"mayz_export_{start}_{end}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

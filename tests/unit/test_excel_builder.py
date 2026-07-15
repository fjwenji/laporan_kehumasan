"""
Unit Tests for Excel Builder Module (src/excel_builder.py)
Test Case IDs: EB-001 to EB-005

Metodologi: Speech-Driven Development (SDD)
Format: Given-When-Then (GWT)
"""

import pytest
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from src.excel_builder import (
    build_output,
    save_output,
    style_header,
    style_body,
    set_widths,
)
from src.parser import AccountRow, ScrapeRow


# ============================================================
# Test Case: EB-001 - Build Output Excel
# ============================================================

class TestBuildOutput:
    """EB-001: System generates Excel with correct headers and data"""

    @pytest.fixture
    def sample_accounts(self):
        """Sample accounts for testing."""
        return [
            AccountRow(
                no=1,
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                manual_judul="",
                manual_link="",
                manual_reach="",
                agenda_no="A001",
                agenda_topic="Topik Jakarta",
            ),
            AccountRow(
                no=2,
                nama_kanwil="Kanwil DJPb Bandung",
                url_akun="https://www.instagram.com/djpbbandung/",
                manual_judul="Judul Manual Bandung",
                manual_link="https://instagram.com/p/TEST/",
                manual_reach="500",
                agenda_no="A002",
                agenda_topic="Topik Bandung",
            ),
        ]

    @pytest.fixture
    def sample_rows(self):
        """Sample scrape rows for testing."""
        return [
            ScrapeRow(
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                post_url="https://www.instagram.com/p/JKT001/",
                shortcode="JKT001",
                tanggal_postingan=datetime(2026, 6, 15, 10, 30),
                media_type="Post / Picture / Carousel",
                caption="Kegiatan monitoring di Jakarta",
                like_count=1500,
                comment_count=120,
                total_engagement=1620,
                status_periode="Masuk Periode",
                status_scraping="FULL_SUCCESS",
                catatan="",
            ),
            ScrapeRow(
                nama_kanwil="Kanwil DJPb Bandung",
                url_akun="https://www.instagram.com/djpbbandung/",
                post_url="https://www.instagram.com/p/BDG001/",
                shortcode="BDG001",
                tanggal_postingan=datetime(2026, 6, 16, 14, 0),
                media_type="Reels",
                caption="Publikasi kegiatan Bandung",
                like_count=800,
                comment_count=45,
                total_engagement=845,
                status_periode="Masuk Periode",
                status_scraping="FULL_SUCCESS",
                catatan="",
            ),
        ]

    def test_build_output_creates_workbook(self, sample_accounts, sample_rows):
        """
        Scenario: System generates Excel workbook

        Given accounts: 2 Kanwil
        And rows: 2 ScrapeRow
        And selected_fields: ["Like Count", "Comment Count"]

        When saya memanggil build_output(accounts, rows, selected_fields)

        Then workbook bytes dikembalikan
        """
        result = build_output(
            accounts=sample_accounts,
            rows=sample_rows,
            selected_fields=["Like Count", "Comment Count"],
            only_period=False,
            include_raw=False,
        )

        assert result is not None
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_build_output_sheet_name(self, sample_accounts, sample_rows):
        """
        Scenario: Output sheet is named "DJPb"

        When saya memanggil build_output

        Then sheet aktif bernama "DJPb"
        """
        result = build_output(
            accounts=sample_accounts,
            rows=sample_rows,
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        assert "DJPb" in wb.sheetnames

    def test_build_output_headers_present(self, sample_accounts, sample_rows):
        """
        Scenario: Headers are in row 6

        When saya memanggil build_output

        Then row 6 berisi header: No., Nama Kanwil, ...
        """
        result = build_output(
            accounts=sample_accounts,
            rows=sample_rows,
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Header should be in row 6
        headers = [ws.cell(row=6, column=c).value for c in range(1, 12)]
        assert "No." in headers
        assert "Nama Kanwil" in headers
        assert "Tanggal Postingan" in headers

    def test_build_output_extra_fields_added(self, sample_accounts, sample_rows):
        """
        Scenario: Extra fields are added after base headers

        Given selected_fields: ["Like Count", "Comment Count"]

        When saya memanggil build_output

        Then header termasuk kolom tambahan di posisi yang benar
        """
        result = build_output(
            accounts=sample_accounts,
            rows=sample_rows,
            selected_fields=["Like Count", "Comment Count"],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Get all headers from row 6
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=6, column=col).value
            if val:
                headers.append(val)

        assert "Like Count" in headers
        assert "Comment Count" in headers


# ============================================================
# Test Case: EB-002 - Include Raw Sheet Option
# ============================================================

class TestIncludeRawSheet:
    """EB-002: System optionally includes raw scraping data sheet"""

    @pytest.fixture
    def sample_accounts(self):
        return [
            AccountRow(
                no=1,
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                manual_judul="",
                manual_link="",
                manual_reach="",
                agenda_no="A001",
                agenda_topic="Topik",
            ),
        ]

    @pytest.fixture
    def sample_rows(self):
        return [
            ScrapeRow(
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                post_url="https://www.instagram.com/p/JKT001/",
                shortcode="JKT001",
                tanggal_postingan=datetime(2026, 6, 15),
                media_type="Post",
                caption="Test caption",
                like_count=100,
                comment_count=10,
                total_engagement=110,
                status_periode="Masuk Periode",
                status_scraping="FULL_SUCCESS",
                catatan="",
            ),
        ]

    def test_include_raw_creates_second_sheet(self, sample_accounts, sample_rows):
        """
        Scenario: include_raw=True creates second sheet

        Given include_raw = true

        When saya memanggil build_output(..., include_raw=True)

        Then workbook memiliki 2 sheets:
          - "DJPb" (formatted output)
          - "Raw_Scraping" (raw data)
        """
        result = build_output(
            accounts=sample_accounts,
            rows=sample_rows,
            selected_fields=[],
            only_period=False,
            include_raw=True,
        )

        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) == 2
        assert "DJPb" in wb.sheetnames
        assert "Raw_Scraping" in wb.sheetnames

    def test_include_raw_false_single_sheet(self, sample_accounts, sample_rows):
        """
        Scenario: include_raw=False creates only output sheet

        Given include_raw = false

        When saya memanggil build_output(..., include_raw=False)

        Then workbook memiliki 1 sheet saja
        """
        result = build_output(
            accounts=sample_accounts,
            rows=sample_rows,
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) == 1

    def test_raw_sheet_has_all_fields(self, sample_accounts, sample_rows):
        """
        Scenario: Raw sheet contains all scraping fields

        When saya memanggil build_output dengan include_raw=True

        Then raw sheet memiliki kolom: shortcode, caption, like_count, etc.
        """
        result = build_output(
            accounts=sample_accounts,
            rows=sample_rows,
            selected_fields=[],
            only_period=False,
            include_raw=True,
        )

        wb = load_workbook(BytesIO(result))
        raw_ws = wb["Raw_Scraping"]

        # Check headers (row 1)
        headers = [raw_ws.cell(row=1, column=c).value for c in range(1, 15)]
        assert "shortcode" in headers
        assert "caption" in headers
        assert "like_count" in headers
        assert "comment_count" in headers


# ============================================================
# Test Case: EB-003 - Hyperlink pada Link Column
# ============================================================

class TestHyperlink:
    """EB-003: System adds hyperlink to post URL column"""

    @pytest.fixture
    def single_account(self):
        return [
            AccountRow(
                no=1,
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                manual_judul="",
                manual_link="",
                manual_reach="",
                agenda_no="A001",
                agenda_topic="Topik",
            ),
        ]

    @pytest.fixture
    def single_row(self):
        return [
            ScrapeRow(
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                post_url="https://www.instagram.com/p/JKT001/",
                shortcode="JKT001",
                tanggal_postingan=datetime(2026, 6, 15),
                media_type="Post",
                caption="Test",
                like_count=100,
                comment_count=10,
                total_engagement=110,
                status_periode="Masuk Periode",
                status_scraping="FULL_SUCCESS",
                catatan="",
            ),
        ]

    def test_hyperlink_added_to_link_column(self, single_account, single_row):
        """
        Scenario: Hyperlink is added to column G (Link)

        Given ScrapeRow dengan post_url = "https://instagram.com/p/ABC123/"

        When saya memanggil build_output

        Then cell G9 memiliki hyperlink ke post_url
        """
        result = build_output(
            accounts=single_account,
            rows=single_row,
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Column G is the Link column (position 7)
        link_cell = ws.cell(row=9, column=7)
        assert link_cell.hyperlink is not None
        assert link_cell.hyperlink.target == "https://www.instagram.com/p/JKT001/"

    def test_hyperlink_style_is_hyperlink(self, single_account, single_row):
        """
        Scenario: Link cell has Hyperlink style

        When saya memanggil build_output

        Then cell G9 style = "Hyperlink"
        """
        result = build_output(
            accounts=single_account,
            rows=single_row,
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        link_cell = ws.cell(row=9, column=7)
        assert link_cell.style == "Hyperlink"


# ============================================================
# Test Case: EB-004 - Filter by Period
# ============================================================

class TestFilterByPeriod:
    """EB-004: System filters rows by period when only_period=True"""

    @pytest.fixture
    def mixed_period_account(self):
        return [
            AccountRow(
                no=1,
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                manual_judul="",
                manual_link="",
                manual_reach="",
                agenda_no="A001",
                agenda_topic="Topik",
            ),
        ]

    @pytest.fixture
    def mixed_period_rows(self):
        return [
            ScrapeRow(
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                post_url="https://www.instagram.com/p/JKT001/",
                shortcode="JKT001",
                tanggal_postingan=datetime(2026, 6, 15),
                media_type="Post",
                caption="Within period",
                like_count=100,
                comment_count=10,
                total_engagement=110,
                status_periode="Masuk Periode",
                status_scraping="FULL_SUCCESS",
                catatan="",
            ),
            ScrapeRow(
                nama_kanwil="Kanwil DJPb Jakarta",
                url_akun="https://www.instagram.com/djpbjakarta/",
                post_url="https://www.instagram.com/p/JKT002/",
                shortcode="JKT002",
                tanggal_postingan=datetime(2026, 5, 15),
                media_type="Post",
                caption="Outside period",
                like_count=200,
                comment_count=20,
                total_engagement=220,
                status_periode="Di Luar Periode",
                status_scraping="FULL_SUCCESS",
                catatan="",
            ),
        ]

    def test_only_period_true_filters(self, mixed_period_account, mixed_period_rows):
        """
        Scenario: only_period=True filters out non-period rows

        Given rows:
          - Row A: status_periode = "Masuk Periode"
          - Row B: status_periode = "Di Luar Periode"

        When saya memanggil build_output(..., only_period=True)

        Then hanya Row A yang masuk Excel
        """
        result = build_output(
            accounts=mixed_period_account,
            rows=mixed_period_rows,
            selected_fields=[],
            only_period=True,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Count data rows (starting from row 9)
        data_count = 0
        for row in range(9, ws.max_row + 1):
            if ws.cell(row=row, column=7).value:  # Link column has URL
                data_count += 1

        assert data_count == 1

    def test_only_period_false_includes_all(self, mixed_period_account, mixed_period_rows):
        """
        Scenario: only_period=False includes all rows

        When saya memanggil build_output(..., only_period=False)

        Then semua rows masuk Excel
        """
        result = build_output(
            accounts=mixed_period_account,
            rows=mixed_period_rows,
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Count data rows
        data_count = 0
        for row in range(9, ws.max_row + 1):
            if ws.cell(row=row, column=7).value:
                data_count += 1

        assert data_count == 2


# ============================================================
# Additional Excel Builder Tests
# ============================================================

class TestSaveOutput:
    """Tests for save_output function"""

    def test_save_output_creates_file(self, tmp_path, sample_account_row, sample_scrape_row):
        """
        Scenario: System saves output to file

        When saya memanggil save_output(content, export_dir)

        Then file .xlsx dibuat di export_dir
        """
        content = build_output(
            accounts=[sample_account_row],
            rows=[sample_scrape_row],
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        result = save_output(content, tmp_path)

        assert result.exists()
        assert result.suffix == ".xlsx"
        assert "Pelaporan" in result.name

    def test_save_output_filename_timestamp(self, tmp_path, sample_account_row, sample_scrape_row):
        """
        Scenario: Output filename includes timestamp

        When saya memanggil save_output

        Then filename format: Pelaporan_Juni_2026_output_YYYYMMDD_HHMMSS.xlsx
        """
        content = build_output(
            accounts=[sample_account_row],
            rows=[sample_scrape_row],
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        result = save_output(content, tmp_path)

        # Check filename pattern
        assert "Pelaporan" in result.name
        assert "_output_" in result.name
        assert result.name.endswith(".xlsx")


class TestTitleAndHeader:
    """Tests for Excel title and header formatting"""

    def test_title_rows_present(self, sample_account_row, sample_scrape_row):
        """
        Scenario: Excel has title rows at top

        When saya memanggil build_output

        Then:
          - Row 1: "DAFTAR LINK PELAPORAN EKSIS VERTIKAL"
          - Row 2: "DIREKTORAT JENDERAL PERBENDAHARAAN"
        """
        content = build_output(
            accounts=[sample_account_row],
            rows=[sample_scrape_row],
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(content))
        ws = wb.active

        assert ws["A1"].value == "DAFTAR LINK PELAPORAN EKSIS VERTIKAL"
        assert ws["A2"].value == "DIREKTORAT JENDERAL PERBENDAHARAAN"

    def test_header_row_styling(self, sample_account_row, sample_scrape_row):
        """
        Scenario: Header row has blue background and white text

        When saya memanggil build_output

        Then header row styling applied
        """
        content = build_output(
            accounts=[sample_account_row],
            rows=[sample_scrape_row],
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(content))
        ws = wb.active

        header_cell = ws.cell(row=6, column=1)
        # fgColor.rgb includes alpha channel (8 hex digits)
        fill_rgb = header_cell.fill.fgColor.rgb
        assert fill_rgb.endswith("1F4E78"), f"Expected fill color ending with 1F4E78, got {fill_rgb}"
        # Font color also includes alpha channel
        font_rgb = header_cell.font.color.rgb
        assert font_rgb.endswith("FFFFFF"), f"Expected font color ending with FFFFFF, got {font_rgb}"


class TestDataRowValues:
    """Tests for data row value mapping"""

    def test_caption_fallback_to_manual(self, sample_account_row):
        """
        Scenario: Empty caption falls back to manual_judul

        Given ScrapeRow dengan caption = ""
        And AccountRow dengan manual_judul = "Judul Manual"

        When saya memanggil build_output

        Then kolom "Judul Postingan" = "Judul Manual"
        """
        account = sample_account_row
        row = ScrapeRow(
            nama_kanwil=account.nama_kanwil,
            url_akun=account.url_akun,
            post_url="https://www.instagram.com/p/TEST/",
            shortcode="TEST",
            tanggal_postingan=datetime(2026, 6, 15),
            media_type="Post",
            caption="",  # Empty caption
            like_count=0,
            comment_count=0,
            total_engagement=0,
            status_periode="Masuk Periode",
            status_scraping="FULL_SUCCESS",
            catatan="",
        )

        content = build_output(
            accounts=[account],
            rows=[row],
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(content))
        ws = wb.active

        # Column F (Judul Postingan) should have manual_judul
        judul_cell = ws.cell(row=9, column=6)
        assert judul_cell.value == "Judul Manual Test"

    def test_no_data_message(self):
        """
        Scenario: Empty result shows message

        Given empty accounts and rows

        When saya memanggil build_output

        Then row 9 contains "Tidak ada data postingan"
        """
        content = build_output(
            accounts=[],
            rows=[],
            selected_fields=[],
            only_period=False,
            include_raw=False,
        )

        wb = load_workbook(BytesIO(content))
        ws = wb.active

        assert "Tidak ada data postingan" in str(ws.cell(row=9, column=1).value)
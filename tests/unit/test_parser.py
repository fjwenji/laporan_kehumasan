"""
Unit Tests for Parser Module (src/parser.py)
Test Case IDs: P-001 to P-007

Metodologi: Speech-Driven Development (SDD)
Format: Given-When-Then (GWT)
"""

import pytest
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook

from src.parser import (
    AccountRow,
    ScrapeRow,
    safe_text,
    normalize_url,
    extract_shortcode,
    detect_media_type,
    parse_dt_to_wib,
    parse_number_token,
    parse_engagement_from_text,
    extract_caption_from_meta,
    status_periode,
    load_accounts_from_excel,
)


# ============================================================
# Test Case: P-001 - Load Akun dari Excel
# ============================================================

class TestLoadAccountsFromExcel:
    """P-001: Load accounts from valid Excel template"""

    def test_load_accounts_from_excel_valid(self):
        """
        Scenario: User uploads valid Excel template with Instagram links

        Given file Excel berisi:
          | Kolom A      | Kolom C                    |
          | Kanwil DJPb Jakarta | https://www.instagram.com/djpbjakarta |

        When saya memanggil load_accounts_from_excel(file_bytes)

        Then sistem mengembalikan list AccountRow dengan:
          - nama_kanwil = "Kanwil DJPb Jakarta"
          - url_akun = "https://www.instagram.com/djpbjakarta/"
        """
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "DJPb"

        # Header
        headers = ["Nama Kanwil", "", "URL Instagram", "", "", "Manual Judul", "Manual Link", "", "Manual Reach", "Agenda No", "Topik"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Data
        ws.cell(row=2, column=1, value="Kanwil DJPb Jakarta")
        ws.cell(row=2, column=3, value="https://www.instagram.com/djpbjakarta/")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_bytes = output.getvalue()

        # Act
        accounts = load_accounts_from_excel(file_bytes)

        # Assert
        assert len(accounts) == 1
        assert accounts[0].nama_kanwil == "Kanwil DJPb Jakarta"
        assert "djpbjakarta" in accounts[0].url_akun
        assert accounts[0].url_akun == "https://www.instagram.com/djpbjakarta/"

    def test_load_accounts_case_insensitive_sheet(self):
        """
        P-002: System auto-detects DJPb sheet regardless of name casing

        Given workbook dengan sheet bernama "djpb" (lowercase)

        When saya memanggil load_accounts_from_excel

        Then sheet "djpb" digunakan untuk parsing
        """
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "djpb"  # lowercase

        ws.cell(row=1, column=1, value="Kanwil DJPb Test")
        ws.cell(row=1, column=3, value="https://www.instagram.com/testaccount/")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        accounts = load_accounts_from_excel(output.getvalue())

        # Assert
        assert len(accounts) == 1

    def test_ignore_rows_without_instagram(self):
        """
        P-003: System ignores rows without valid Instagram URLs

        Given file Excel dengan baris:
          | Row 1 | https://www.instagram.com/validaccount |
          | Row 2 | https://facebook.com/notinstagram     |
          | Row 3 | (kosong)                              |

        When saya memanggil load_accounts_from_excel

        Then hanya 1 account yang dikembalikan
        """
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "DJPb"

        ws.cell(row=1, column=1, value="Valid Account")
        ws.cell(row=1, column=3, value="https://www.instagram.com/validaccount/")

        ws.cell(row=2, column=1, value="Facebook Account")
        ws.cell(row=2, column=3, value="https://facebook.com/notinstagram")

        ws.cell(row=3, column=1, value="Empty Row")
        ws.cell(row=3, column=3, value="")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        accounts = load_accounts_from_excel(output.getvalue())

        # Assert
        assert len(accounts) == 1
        assert "validaccount" in accounts[0].url_akun

    def test_deduplicate_accounts(self):
        """
        P-004: System removes duplicate Instagram accounts

        Given file Excel dengan URL duplikat:
          | https://www.instagram.com/djpbjakarta |
          | https://www.instagram.com/DJPBJAKARTA |

        When saya memanggil load_accounts_from_excel

        Then hanya 1 account yang dikembalikan
        """
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "DJPb"

        ws.cell(row=1, column=1, value="Kanwil 1")
        ws.cell(row=1, column=3, value="https://www.instagram.com/djpbjakarta/")

        ws.cell(row=2, column=1, value="Kanwil 1 Duplicate")
        ws.cell(row=2, column=3, value="https://www.instagram.com/DJPBJAKARTA/")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        accounts = load_accounts_from_excel(output.getvalue())

        # Assert
        assert len(accounts) == 1

    def test_ignore_post_urls_not_profile(self):
        """
        Scenario: System ignores individual post URLs, only processes profile URLs

        Given URL: https://www.instagram.com/p/ABC123/

        When saya memanggil load_accounts_from_excel

        Then ValueError raised (tidak ada akun valid ditemukan)
        """
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "DJPb"

        ws.cell(row=1, column=1, value="Post URL Test")
        ws.cell(row=1, column=3, value="https://www.instagram.com/p/ABC123/")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act & Assert
        # Post URLs tidak diproses, jadi tidak ada account valid
        # -> ValueError raised karena tidak ada akun ditemukan
        with pytest.raises(ValueError, match="Tidak ada akun Instagram"):
            load_accounts_from_excel(output.getvalue())

    def test_raise_error_when_no_accounts_found(self):
        """
        Scenario: System raises error when no valid accounts found

        Given file Excel tanpa URL Instagram valid

        When saya memanggil load_accounts_from_excel

        Then ValueError raised dengan pesan yang sesuai
        """
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "DJPb"

        ws.cell(row=1, column=1, value="No Instagram")
        ws.cell(row=1, column=3, value="https://example.com/")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act & Assert
        with pytest.raises(ValueError, match="Tidak ada akun Instagram"):
            load_accounts_from_excel(output.getvalue())


# ============================================================
# Test Case: P-005 - Extract Shortcode
# ============================================================

class TestExtractShortcode:
    """P-005: Extract shortcode from various Instagram URL formats"""

    @pytest.mark.parametrize("url,expected", [
        ("https://www.instagram.com/p/ABC123DEF/", "ABC123DEF"),
        ("https://instagram.com/reel/GHI789/", "GHI789"),
        ("https://www.instagram.com/tv/JKL012/", "JKL012"),
        ("https://www.instagram.com/p/ABC123DEF", "ABC123DEF"),  # tanpa trailing slash
        ("/p/MXN123/", "MXN123"),  # relative URL
        ("", ""),  # empty
        ("https://www.instagram.com/djpbjakarta/", ""),  # profile URL, not post
    ])
    def test_extract_shortcode_various_formats(self, url, expected):
        """
        Given URLs berikut:
          | URL Format                                      | Expected Shortcode |
          | https://www.instagram.com/p/ABC123DEF/          | ABC123DEF         |
          | https://instagram.com/reel/GHI789/              | GHI789            |
          | https://www.instagram.com/tv/JKL012/           | JKL012            |

        When saya memanggil extract_shortcode(url)

        Then shortcode yang dikembalikan sesuai ekspektasi
        """
        # Act
        result = extract_shortcode(url)

        # Assert
        assert result == expected

    def test_extract_shortcode_with_query_params(self):
        """
        Scenario: System handles URL with query parameters

        Given URL: https://www.instagram.com/p/ABC123DEF/?utm_source=ig_web

        When saya memanggil extract_shortcode

        Then shortcode = "ABC123DEF"
        """
        result = extract_shortcode("https://www.instagram.com/p/ABC123DEF/?utm_source=ig_web")
        assert result == "ABC123DEF"


# ============================================================
# Test Case: P-006 - Parse Datetime dengan Timezone
# ============================================================

class TestParseDtToWib:
    """P-006: Convert timezone-aware datetime to naive for Excel"""

    def test_parse_datetime_with_timezone(self):
        """
        Given string datetime: "2026-06-15T14:30:00+07:00"

        When saya memanggil parse_dt_to_wib(raw)

        Then hasil adalah datetime object:
          - tanpa timezone info
          - hour = 14
          - minute = 30
        """
        result = parse_dt_to_wib("2026-06-15T14:30:00+07:00")

        assert result is not None
        assert result.year == 2026
        assert result.month == 6
        assert result.day == 15
        assert result.hour == 14
        assert result.minute == 30
        assert result.tzinfo is None  # Naive datetime

    def test_parse_datetime_with_z_suffix(self):
        """
        Scenario: System handles ISO format with Z suffix

        Given: "2026-06-15T10:30:00Z"

        When: saya memanggil parse_dt_to_wib

        Then: timezone stripped, tetap naive
        """
        result = parse_dt_to_wib("2026-06-15T10:30:00Z")

        assert result is not None
        assert result.tzinfo is None

    def test_parse_datetime_invalid_returns_none(self):
        """
        Scenario: System returns None for invalid datetime

        Given: "invalid-date-string"

        When: saya memanggil parse_dt_to_wib

        Then: returns None
        """
        result = parse_dt_to_wib("invalid-date-string")
        assert result is None

    def test_parse_datetime_empty_returns_none(self):
        """
        Scenario: System returns None for empty string

        Given: ""

        When: saya memanggil parse_dt_to_wib

        Then: returns None
        """
        result = parse_dt_to_wib("")
        assert result is None


# ============================================================
# Test Case: P-007 - Parse Number dengan Suffix
# ============================================================

class TestParseNumberToken:
    """P-007: Parse numbers with K, M, rb, jt suffixes"""

    @pytest.mark.parametrize("token,expected", [
        ("1.5k", 1500),
        ("2.3K", 2300),
        ("100rb", 100000),
        ("1.2jt", 1200000),
        ("3M", 3000000),
        ("500", 500),
        ("1k", 1000),
        ("10jt", 10000000),
    ])
    def test_parse_number_with_suffix(self, token, expected):
        """
        Given tokens berikut:
          | Token | Expected Result |
          | "1.5k" | 1500           |
          | "2.3K" | 2300           |
          | "100rb" | 100000        |
          | "1.2jt" | 1200000       |
          | "3M"    | 3000000       |

        When saya memanggil parse_number_token(token)

        Then hasil sesuai ekspektasi
        """
        result = parse_number_token(token)
        assert result == expected

    def test_parse_number_with_space(self):
        """
        Scenario: System handles number with spaces

        Given: "1.5 k"

        When: saya memanggil parse_number_token

        Then: result = 1500
        """
        result = parse_number_token("1.5 k")
        assert result == 1500

    def test_parse_number_empty_returns_none(self):
        """
        Scenario: System returns None for empty token

        Given: ""

        When: saya memanggil parse_number_token

        Then: returns None
        """
        result = parse_number_token("")
        assert result is None


# ============================================================
# Additional Parser Tests
# ============================================================

class TestDetectMediaType:
    """Tests for detect_media_type function"""

    @pytest.mark.parametrize("url,expected", [
        ("https://www.instagram.com/reel/ABC123/", "reels"),
        ("https://www.instagram.com/p/ABC123/", "image"),
        ("https://www.instagram.com/tv/ABC123/", "video"),
        ("https://example.com/", "unknown"),
    ])
    def test_detect_media_type(self, url, expected):
        result = detect_media_type(url)
        assert result == expected


class TestParseEngagementFromText:
    """Tests for parse_engagement_from_text function"""

    def test_parse_engagement_likes_and_comments(self):
        """
        Scenario: System extracts like and comment counts from text

        Given: "Post with 1,234 likes and 56 comments"

        When: saya memanggil parse_engagement_from_text

        Then: (1234, 56)
        """
        text = "Post with 1,234 likes and 56 comments"
        likes, comments = parse_engagement_from_text(text)

        assert likes == 1234
        assert comments == 56

    def test_parse_engagement_with_suffix(self):
        """
        Scenario: System handles K suffix in engagement text

        Given: "This post has 5.2K likes"

        When: saya memanggil parse_engagement_from_text

        Then: likes = 5200
        """
        text = "This post has 5.2K likes"
        likes, comments = parse_engagement_from_text(text)

        assert likes == 5200


class TestStatusPeriode:
    """Tests for status_periode function"""

    def test_within_period(self):
        """
        Scenario: Post within period returns "Masuk Periode"

        Given:
          - tanggal: 2026-06-10
          - period_start: 2026-06-01
          - period_end: 2026-06-30

        When: saya memanggil status_periode

        Then: "Masuk Periode"
        """
        tanggal = datetime(2026, 6, 10)
        period_start = datetime(2026, 6, 1)
        period_end = datetime(2026, 6, 30)

        result = status_periode(tanggal, period_start, period_end)
        assert result == "Masuk Periode"

    def test_outside_period(self):
        """
        Scenario: Post outside period returns "Di Luar Periode"

        Given:
          - tanggal: 2026-07-15
          - period_start: 2026-06-01
          - period_end: 2026-06-30

        When: saya memanggil status_periode

        Then: "Di Luar Periode"
        """
        tanggal = datetime(2026, 7, 15)
        period_start = datetime(2026, 6, 1)
        period_end = datetime(2026, 6, 30)

        result = status_periode(tanggal, period_start, period_end)
        assert result == "Di Luar Periode"

    def test_null_tanggal_returns_need_review(self):
        """
        Scenario: Null tanggal returns "Perlu Cek Manual"

        Given: tanggal = None

        When: saya memanggil status_periode

        Then: "Perlu Cek Manual"
        """
        result = status_periode(None, datetime(2026, 6, 1), datetime(2026, 6, 30))
        assert result == "Perlu Cek Manual"


class TestSafeText:
    """Tests for safe_text utility function"""

    def test_safe_text_none(self):
        """Given: None input -> returns empty string"""
        assert safe_text(None) == ""

    def test_safe_text_strips_whitespace(self):
        """Given: "  text  " -> returns "text" """
        assert safe_text("  text  ") == "text"

    def test_safe_text_converts_int(self):
        """Given: 123 -> returns "123" """
        assert safe_text(123) == "123"


class TestNormalizeUrl:
    """Tests for normalize_url function"""

    def test_normalize_url_adds_https(self):
        """Given: "www.instagram.com/user" -> returns "https://www.instagram.com/user/" """
        result = normalize_url("www.instagram.com/user")
        assert result.startswith("https://")

    def test_normalize_url_preserves_https(self):
        """Given: "https://www.instagram.com/user" -> unchanged"""
        result = normalize_url("https://www.instagram.com/user")
        assert result == "https://www.instagram.com/user"


# ============================================================
# Integration-like Tests (Parser + Excel Builder)
# ============================================================

class TestAccountRowToScrapeRow:
    """Tests for AccountRow to ScrapeRow conversion"""

    def test_account_row_fields(self, sample_account_row):
        """
        Scenario: AccountRow contains all expected fields

        When: saya membuat AccountRow

        Then semua field terisi dengan benar
        """
        assert sample_account_row.no == 1
        assert sample_account_row.nama_kanwil == "Kanwil DJPb Jakarta"
        assert "djpbjakarta" in sample_account_row.url_akun
        assert sample_account_row.manual_judul == "Judul Manual Test"
        assert sample_account_row.agenda_no == "A001"

    def test_scrape_row_fields(self, sample_scrape_row):
        """
        Scenario: ScrapeRow contains all expected fields

        When: saya membuat ScrapeRow

        Then semua field terisi dengan benar
        """
        assert sample_scrape_row.nama_kanwil == "Kanwil DJPb Jakarta"
        assert sample_scrape_row.post_url == "https://www.instagram.com/p/ABC123DEF/"
        assert sample_scrape_row.shortcode == "ABC123DEF"
        assert sample_scrape_row.tanggal_postingan == datetime(2026, 6, 15, 10, 30)
        assert sample_scrape_row.like_count == 1500
        assert sample_scrape_row.total_engagement == 1620
        assert sample_scrape_row.status_periode == "Masuk Periode"
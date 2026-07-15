"""
Pytest Configuration and Fixtures for Mayz DJPb Tests.
"""

import os
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))
os.environ["TESTING"] = "true"
@pytest.fixture
def sample_excel_bytes():
    """Generate sample Excel file bytes for testing."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DJPb"
    headers = ["Nama Kanwil", "Kolom2", "URL Instagram", "Kolom4", "Kolom5", "Manual Judul", "Manual Link", "Kolom8", "Manual Reach", "Agenda No", "Topik"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    data = [
        ["Kanwil DJPb Jakarta", "Data2", "https://www.instagram.com/djpbjakarta/", "Data4", "Data5", "Judul Manual 1", "https://instagram.com/p/TEST123/", "Data8", "Reach 1", "A001", "Topik Agenda 1"],
        ["Kanwil DJPb Bandung", "Data2", "https://www.instagram.com/djpbbandung/", "Data4", "Data5", "Judul Manual 2", "", "Data8", "Reach 2", "A002", "Topik Agenda 2"],
        ["Kanwil DJPb Surabaya", "Data2", "https://www.instagram.com/djpbsurabaya/", "Data4", "Data5", "", "", "Data8", "", "", ""],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

@pytest.fixture
def sample_account_row():
    """Sample AccountRow for testing."""
    from src.parser import AccountRow

    return AccountRow(
        no=1,
        nama_kanwil="Kanwil DJPb Jakarta",
        url_akun="https://www.instagram.com/djpbjakarta/",
        manual_judul="Judul Manual Test",
        manual_link="https://instagram.com/p/TEST123/",
        manual_reach="1000",
        agenda_no="A001",
        agenda_topic="Topik Test",
    )
@pytest.fixture
def sample_scrape_row():
    """Sample ScrapeRow for testing."""
    from src.parser import ScrapeRow

    return ScrapeRow(
        nama_kanwil="Kanwil DJPb Jakarta",
        url_akun="https://www.instagram.com/djpbjakarta/",
        post_url="https://www.instagram.com/p/ABC123DEF/",
        shortcode="ABC123DEF",
        tanggal_postingan=datetime(2026, 6, 15, 10, 30),
        media_type="Post / Picture / Carousel",
        caption="Detail kegiatan monitoring hari ini di Jakarta",
        like_count=1500,
        comment_count=120,
        total_engagement=1620,
        status_periode="Masuk Periode",
        status_scraping="FULL_SUCCESS",
        catatan="",
    )
@pytest.fixture
def sample_html():
    """Sample Instagram post HTML for testing."""
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <meta property="og:description" content="5,432 likes, 123 comments - djpbjakarta on June 15, 2026: Detail kegiatan monitoring publikasi di Jakarta">
        <meta property="og:url" content="https://www.instagram.com/p/ABC123DEF/">
        <meta property="og:image" content="https://example.com/image.jpg">
        <link rel="canonical" href="https://www.instagram.com/p/ABC123DEF/">
    </head>
    <body>
        <article>
            <time datetime="2026-06-15T10:30:00+07:00"></time>
            <div class="_aagv">
                <img src="https://example.com/image.jpg" alt="Foto kegiatan monitoring">
            </div>
        </article>
    </body>
    </html>
    """
@pytest.fixture
def mock_page():
    """Mock Playwright page for extraction testing."""
    class MockPage:
        def __init__(self):
            self.url = "https://www.instagram.com/p/ABC123DEF/"
            self._html = ""
        @property
        def content(self):
            return self._html
        def locator(self, selector):
            return MockLocator(selector)
        def evaluate(self, script):
            if "querySelector" in script:
                if "og:description" in script:
                    return "5,432 likes, 123 comments - djpbjakarta on June 15, 2026: Detail kegiatan"
                if "og:url" in script:
                    return "https://www.instagram.com/p/ABC123DEF/"
                if "canonical" in script:
                    return "https://www.instagram.com/p/ABC123DEF/"
            return ""
    class MockLocator:
        def __init__(self, selector):
            self.selector = selector
            self.count_value = 0
            self.first = self
        def count(self):
            return self.count_value
        def all(self):
            return []
        def first(self):
            return MockElement(self.selector)
        def get_attribute(self, attr, timeout=2000):
            if self.selector == "meta[property='og:description']":
                if attr == "content":
                    return "5,432 likes, 123 comments - djpbjakarta on June 15, 2026: Detail kegiatan"
            return None
        def inner_text(self, timeout=2000):
            return ""
    class MockElement:
        def __init__(self, selector):
            self.selector = selector
        def get_attribute(self, attr, timeout=2000):
            return None
        def inner_text(self, timeout=2000):
            return ""
    return MockPage()
@pytest.fixture
def test_db_config():
    """Test database configuration."""
    return {
        "host": os.getenv("TEST_MYSQL_HOST", "localhost"),
        "port": int(os.getenv("TEST_MYSQL_PORT", "3306")),
        "user": os.getenv("TEST_MYSQL_USER", "root"),
        "password": os.getenv("TEST_MYSQL_PASSWORD", ""),
        "database": os.getenv("TEST_MYSQL_DATABASE", "mayz_monitoring_test"),
    }
@pytest.fixture(autouse=True)
def setup_test_env(monkeypatch, tmp_path):
    """Setup test environment for all tests."""
    data_dir = tmp_path / "data"
    export_dir = tmp_path / "exports"
    data_dir.mkdir()
    export_dir.mkdir()
    monkeypatch.setattr("src.config.DATA_DIR", data_dir)
    monkeypatch.setattr("src.config.EXPORT_DIR", export_dir)
def create_test_excel(rows_data, sheet_name="DJPb"):
    """Helper to create test Excel file with given rows."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row_idx, row_data in enumerate(rows_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()